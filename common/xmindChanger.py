# mxind数据转换器
import xmind
import json
import re
import uuid
from xmind.core.markerref import MarkerId
from utils.logs import LogManager
from loguru import logger
from fileProcessor import fileProcessor
import pandas as pd
from openpyxl import Workbook, load_workbook
import os
from common.dataProcessor import WriteInfo

class MxindDataProcessor:#  xmind数据整理
    def __init__(self):
        self.xmind_file = r"C:\Users\admin\Desktop\demo.xmind"
        self.logging = LogManager()
        self.case_dict = {}
    def xmind_to_json(self):
        workbook = xmind.load(self.xmind_file)
        sheet = workbook.getData()
        logger.debug("调试信息")
        if sheet:
            logger.debug("调试信息")
            return sheet[0]

class AdvancedTestCaseExtractor: #测试用例数据提取

    def __init__(self, max_depth=None):
        self.max_depth = max_depth
        self.results = []

    def extract(self, json_obj):
        """
        提取测试用例数据
        """
        self.results = []
        root_topic = json_obj.get('topic', {})

        if root_topic:
            # 从根节点开始遍历，传递路径信息
            self._traverse_with_path(root_topic, [], 1)

        return self.results

    def _traverse_with_path(self, node, path, level):
        """
        带路径信息的遍历方法 - 基于层级位置判断
        """
        if self.max_depth is not None and level > self.max_depth:
            return

        # 更新路径
        current_path = path + [node]

        # 基于层级位置判断节点类型
        # 假设结构：第1层为产品名，第2层为模块名，第3层为测试用例
        if level == 4:  # 测试用例节点位于第3层
            # 创建基础测试用例信息
            test_case_info = self._create_test_case_dict_by_position(node, current_path, level)

            # 获取按步骤拆分后的结果
            split_entries = self._extract_steps_and_results_by_position(node, test_case_info)

            if split_entries:
                # 添加拆分后的条目
                self.results.extend(split_entries)
            else:
                # 如果没有子节点，添加原始条目
                self.results.append(test_case_info)

        # 递归处理子节点
        for child in node.get('topics', []):
            self._traverse_with_path(child, current_path, level + 1)

    def _create_test_case_dict_by_position(self, test_case_node, path, level):
        """
        根据节点在树中的位置创建测试用例字典
        """
        subject = ''
        model = ''

        #  这里处理节点的属性，例如主题、模型等
        if len(path) >= 1:
            subject = path[0].get('title', '')
        if len(path) >= 2:
            model = path[1].get('title', '')
        if len(path) >= 3:
            title = path[2].get('title', '')
        if len(path) >= 4:
            header = path[3].get('title', '')
        return {
            'Subject': subject,
            'model': model,
            'title': title,
            'header': header,
            'steps': '',
            'result': ''
        }

    def _extract_steps_and_results_by_position(self, test_case_node, test_case_info):
        """
        基于位置提取测试步骤和预期结果
        假设第一层子节点是步骤，第二层子节点是对应的结果
        """
        new_entries = []

        # 获取子节点（假设为测试步骤）
        child_nodes = test_case_node.get('topics', [])

        for step_node in child_nodes:
            # 获取步骤节点的子节点（假设为预期结果）
            result_nodes = step_node.get('topics', [])

            if result_nodes:
                # 对每个结果节点创建一个条目
                for result_node in result_nodes:
                    new_entry = {
                        'Subject': test_case_info['Subject'],
                        'model': test_case_info['model'],
                        'title': test_case_info['title'],
                        'header': test_case_info['header'],
                        'steps': step_node.get('title', ''),
                        'result': result_node.get('title', '')
                    }
                    new_entries.append(new_entry)
            else:
                # 如果步骤节点没有结果子节点
                new_entry = {
                    'Subject': test_case_info['Subject'],
                    'model': test_case_info['model'],
                    'title': test_case_info['title'],
                    'header': test_case_info['header'],
                    'steps': step_node.get('title', ''),
                    'result': ''
                }
                new_entries.append(new_entry)

        return new_entries

    def get_statistics(self):
        """
        获取提取结果的统计信息

        Returns:
            dict: 统计信息
        """
        if not self.results:
            return {'total_cases': 0}

        total_steps = len(self.results)  # 每个条目代表一个步骤-结果对
        total_results = len(self.results)

        data = {
            'total_cases': len(self.results),
            'total_steps': total_steps,
            'total_results': total_results
        }

        print(data)

class DicToXlsx: # 字典转换成用例
    def __init__(self):
        self.xlsx_file = r'C:\Users\admin\Desktop\demo.xlsx'
        file_dir = os.path.dirname(self.xlsx_file)
        if file_dir and not os.path.exists(file_dir):  # 只有当目录不存在时才创建
            os.makedirs(file_dir, exist_ok=True)
        json_data = MxindDataProcessor().xmind_to_json()
        extractor = AdvancedTestCaseExtractor(max_depth=10)
        self.data = extractor.extract(json_data)
        statistics = extractor.get_statistics()
        print(f"提取到的数据: {len(self.data)} 条")

    def table_data_processing(self):
        try:
            # 如果文件不存在，则创建新的工作簿并设置表头
            if not os.path.exists(self.xlsx_file):
                wb = Workbook()
                sheet = wb.active
                # 定义中文表头
                headers_chinese = ['主题', '模块', '标题', '前置条件', '步骤', '预期结果']
                for col_num, header in enumerate(headers_chinese, 1):
                    sheet.cell(row=1, column=col_num, value=header)
                # 设置表头行高
                sheet.row_dimensions[1].height = 25
                wb.save(self.xlsx_file)
                print(f"已创建新文件并写入表头: {self.xlsx_file}")

            # 加载现有的工作簿
            wb = load_workbook(self.xlsx_file)
            sheet = wb.active

            # 检查表头是否已经存在，如果不存在则添加
            first_cell = sheet.cell(row=1, column=1).value
            headers_chinese = ['主题', '模块', '标题', '前置条件', '步骤', '预期结果']

            # 检查第一列是否包含期望的表头值之一
            if not first_cell or first_cell not in headers_chinese:
                # 重新写入表头
                for col_num, header in enumerate(headers_chinese, 1):
                    sheet.cell(row=1, column=col_num, value=header)
                print("已写入表头到第一行")

            # 获取现有数据的最后一行
            last_row = sheet.max_row
            # 定义英文键名用于从数据字典中获取值
            headers_english = ['Subject', 'model', 'title', 'header', 'steps', 'result']

            # 遍历数据数组，逐行写入
            for row_idx, data_dict in enumerate(self.data, start=last_row + 1):
                for col_idx, header in enumerate(headers_english, start=1):
                    # 从字典中获取对应键的值，如果键不存在则写入空字符串
                    value = data_dict.get(header, '')
                    sheet.cell(row=row_idx, column=col_idx, value=value)

            # 保存工作簿
            wb.save(self.xlsx_file)
            print(f"数据已成功写入: {self.xlsx_file}")

        except PermissionError:
            print(f"权限错误：无法访问文件 {self.xlsx_file}，请确保文件未被其他程序打开")
        except Exception as e:
            print(f"写入Excel文件时发生错误: {str(e)}")


def _build_xmind_topic(xmind_topic, data):
    """
    递归构建 XMind 主题

    Args:
        xmind_topic: XMind 主题对象
        data: 数据字典或字符串
    """
    if isinstance(data, str):
        xmind_topic.setTitle(data)
    elif isinstance(data, dict):
        # 设置当前主题标题
        title = data.get('title', data.get('name', ''))
        if title:
            xmind_topic.setTitle(title)

        # 处理子主题
        topics = data.get('topics', data.get('children', []))
        if topics and isinstance(topics, list):
            for child_data in topics:
                child_topic = xmind_topic.addSubTopic()
                _build_xmind_topic(child_topic, child_data)

        # 处理备注信息（如果有）
        notes = data.get('notes', data.get('comment', ''))
        if notes:
            xmind_topic.setNote(notes)

        # 处理测试用例的特殊字段
        if 'steps' in data or 'result' in data:
            note_content = ""
            if 'steps' in data and data['steps']:
                note_content += f"步骤：{data['steps']}\n"
            if 'result' in data and data['result']:
                note_content += f"预期结果：{data['result']}\n"
            if note_content:
                existing_note = xmind_topic.getNote()
                if existing_note:
                    xmind_topic.setNote(f"{existing_note}\n{note_content}")
                else:
                    xmind_topic.setNote(note_content)
    elif isinstance(data, list):
        # 如果是列表，第一个元素作为标题，其余作为子主题
        if len(data) > 0:
            xmind_topic.setTitle(str(data[0]))
            for item in data[1:]:
                child_topic = xmind_topic.addSubTopic()
                _build_xmind_topic(child_topic, item)


# 测试点的xmind 转化 json
class XmindPointJson:

    def __init__(self):
        self.xmind_file = r"C:\Users\admin\Desktop\test.xmind"

    def readXmindData(self):
        workbook = xmind.load(self.xmind_file)
        sheet = workbook.getData()
        if not sheet:
            print("模板文件中没有工作表")
            return None
        else:
            JsonData = sheet[0]
            print(JsonData)
        return JsonData

    def extract_test_points_data(self):
        """
        从给定的数据中提取测试点和其子项，生成新的JSON格式
        Args:
            data: 包含原始数据的字典
        Returns:
            dict: 新的JSON格式数据，格式为 {"测试点1": ["功能可靠", "性能安全", "使用方便"], "测试点2": [...]}
        """
        # 创建新的结果字典
        result = {}
        data = self.readXmindData()
        # 遍历主题下的topics获取测试点
        topics = data.get('topic', {}).get('topics', [])

        for topic in topics:
            test_point_title = topic.get('title')  # 测试点标题，如"测试点1"

            # 获取该测试点下的所有子主题
            sub_topics = topic.get('topics', [])
            sub_titles = [sub_topic.get('title') for sub_topic in sub_topics]

            # 将测试点作为键，子项列表作为值添加到结果中
            result[test_point_title] = sub_titles

        print(result)



#将AI给的json数据转化为xmind
class TestcasePointJsonToXmind:
    """JSON 数据转换为 XMind 格式的处理器"""

    DEFAULT_ROOT_TITLE = "逻辑图"
    DEFAULT_SHEET_TITLE = "测试点"
    DEFAULT_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'template.xmind')
    BLANK_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'blank_template.xmind')
    FALLBACK_TEMPLATE_PATH = r"C:\Users\admin\Desktop\test.xmind"
    # 标记是否已经警告过模板残留问题
    _template_warning_shown = False
    def __init__(self, template_file=None):
        """
        初始化转换器

        Args:
            template_file: XMind 模板文件路径，默认使用内置模板
        """
        self.logger = LogManager().get_logger() if hasattr(LogManager(), 'get_logger') else logger
        self._ensure_blank_template_exists()
        self.template_file = template_file or self.BLANK_TEMPLATE_PATH


    def _ensure_blank_template_exists(self):
        """确保空白模板文件存在，如果不存在则自动创建"""
        if os.path.exists(self.BLANK_TEMPLATE_PATH):
            return  # 已存在，无需创建

        try:
            import zipfile

            # 创建一个真正的空白 XMind 文件（XMind 8 格式）
            with zipfile.ZipFile(self.BLANK_TEMPLATE_PATH, 'w', zipfile.ZIP_DEFLATED) as zf:
                # content.xml - 只包含一个空的根主题，没有任何子主题
                content_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="no"?>
    <xmap-content xmlns="urn:xmind:xmap:xmlns:content:2.0" xmlns:fo="http://www.w3.org/1999/XSL/Format" xmlns:svg="http://www.w3.org/2000/svg" xmlns:xhtml="http://www.w3.org/1999/xhtml" xmlns:xlink="http://www.w3.org/1999/xlink" modified-by="Lingma" timestamp="1234567890">
    <sheet id="sheet1" modified-by="Lingma" timestamp="1234567890">
    <topic id="root1" modified-by="Lingma" structure-class="org.xmind.ui.logic.right" timestamp="1234567890">
    <title>Root</title>
    </topic>
    </sheet>
    </xmap-content>'''
                zf.writestr('content.xml', content_xml.encode('utf-8'))

                # meta.xml - 元数据
                meta_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="no"?>
    <meta xmlns="urn:xmind:xmap:xmlns:meta:2.0" version="2.0">
    <Creator>Lingma AI</Creator>
    <Created>2024-01-01T00:00:00.000+0800</Created>
    </meta>'''
                zf.writestr('meta.xml', meta_xml.encode('utf-8'))

            self.logger.info(f"✓ 已自动创建空白模板：{self.BLANK_TEMPLATE_PATH}")
        except Exception as e:
            self.logger.error(f"创建空白模板失败：{str(e)}")
            raise
    def _get_template_file(self, template_file=None):

        # 优先级：用户指定 > 空白模板 > 默认模板 > 备用模板
        if template_file and os.path.exists(template_file):
            return template_file
        elif os.path.exists(self.BLANK_TEMPLATE_PATH):
            return self.BLANK_TEMPLATE_PATH
        elif os.path.exists(self.DEFAULT_TEMPLATE_PATH):
            return self.DEFAULT_TEMPLATE_PATH
        elif os.path.exists(self.FALLBACK_TEMPLATE_PATH):
            return self.FALLBACK_TEMPLATE_PATH
        else:
            raise FileNotFoundError(
                f"未找到 XMind 模板文件。\n"
                f"请确保以下路径之一存在：\n"
                f"1. {self.BLANK_TEMPLATE_PATH}（自动生成）\n"
                f"2. {self.DEFAULT_TEMPLATE_PATH}\n"
                f"3. {self.FALLBACK_TEMPLATE_PATH}"
            )

    def json_to_xmind(self, json_data, output_file, template_file=None):

        try:
            # 获取有效的模板文件
            valid_template = self._get_template_file(template_file)

            # 加载模板工作簿
            workbook = xmind.load(valid_template)
            sheet = workbook.getPrimarySheet()

            if not sheet:
                self.logger.error("模板文件中没有工作表")
                return False

            # 设置工作表标题
            sheet_title = json_data.get('title', self.DEFAULT_SHEET_TITLE) if isinstance(json_data,
             dict) else self.DEFAULT_SHEET_TITLE
            sheet.setTitle(sheet_title)

            # 获取根主题
            root_topic = sheet.getRootTopic()


            self._build_root_content_clean(root_topic, json_data)

            # 确保输出目录存在
            self._ensure_directory_exists(output_file)

            # 保存 XMind 文件
            xmind.save(workbook, output_file)
            self.logger.info(f"XMind 文件已成功生成：{output_file}")
            return True

        except FileNotFoundError as e:
            self.logger.error(f"模板文件未找到：{str(e)}")
            return False
        except Exception as e:
            self.logger.error(f"转换 XMind 文件时发生错误：{str(e)}", exc_info=True)
            return False


    def _build_root_content_clean(self, root_topic, json_data):
        """

        Args:
            root_topic: XMind 根主题对象
            json_data: JSON 数据
        """
        if isinstance(json_data, dict):
            if 'topic' in json_data:
                topic_data = json_data['topic']
                # 先设置根主题标题
                if 'title' in topic_data:
                    root_topic.setTitle(topic_data['title'])
                # 然后递归构建子主题
                _build_xmind_topic(root_topic, topic_data)
            else:
                # 直接作为根主题处理
                if 'title' in json_data:
                    root_topic.setTitle(json_data['title'])
                _build_xmind_topic(root_topic, json_data)
        elif isinstance(json_data, list):
            # 如果是列表，每个元素作为一个子主题
            for item in json_data:
                if isinstance(item, dict):
                    child_topic = root_topic.addSubTopic()
                    _build_xmind_topic(child_topic, item)

    def convert_test_points_to_xmind_format(self, input_data, root_title=DEFAULT_ROOT_TITLE, sheet_title=DEFAULT_SHEET_TITLE):
        """
        将测试点数据转换为 XMind 标准 JSON 格式

        Args:
            input_data: 测试点数据（dict）
            root_title: 根节点标题
            sheet_title: 工作表标题

        Returns:
            dict: XMind 格式的 JSON 数据
        """
        category_topics = []

        # 遍历顶层分类（如"功能测试"、"用户使用场景"）
        for category_name, category_data in input_data.items():
            # 为每个分类创建主题
            category_topic = self._create_category_topic(category_name, category_data)
            category_topics.append(category_topic)

        return self._build_xmind_json_structure(
            category_topics,
            root_title,
            sheet_title
        )

    def _create_category_topic(self, category_name, category_data):
        """
        创建分类主题（如"功能测试"）

        Args:
            category_name: 分类名称
            category_data: 分类下的数据（dict 或 list）

        Returns:
            dict: 分类主题数据
        """
        sub_topics = self._extract_sub_topics(category_data)

        return {
            'id': self._generate_id(),
            'link': None,
            'title': category_name,  # 使用实际分类名，而非"测试点1"
            'note': None,
            'label': None,
            'comment': None,
            'markers': [],
            'topics': sub_topics
        }

    def _extract_sub_topics(self, subcategories):
        """
        提取子主题列表

        Args:
            subcategories: 子类别数据（dict 或 list）

        Returns:
            list: 子主题列表
        """
        sub_topics = []

        if isinstance(subcategories, dict):
            for sub_key, sub_value in subcategories.items():
                # 如果值是列表，创建带子主题的主题节点
                if isinstance(sub_value, list):
                    parent_topic = self._create_sub_topic(sub_key)
                    # 为列表中的每个元素创建子主题
                    child_topics = []
                    for item in sub_value:
                        child_topics.append(self._create_sub_topic(str(item)))
                    parent_topic['topics'] = child_topics
                    sub_topics.append(parent_topic)
                else:
                    # 如果值是字符串或其他类型，直接创建主题
                    sub_topics.append(self._create_sub_topic(sub_key))
        elif isinstance(subcategories, list):
            for item in subcategories:
                sub_topics.append(self._create_sub_topic(str(item)))

        return sub_topics

    def _create_sub_topic(self, title):
        """
        创建单个子主题

        Args:
            title: 子主题标题

        Returns:
            dict: 子主题数据
        """
        return {
            'id': self._generate_id(),
            'link': None,
            'title': title,
            'note': None,
            'label': None,
            'comment': None,
            'markers': []
        }

    def _build_xmind_json_structure(self, test_point_topics, root_title, sheet_title):
        """
        构建完整的 XMind JSON 结构

        Args:
            test_point_topics: 测试点主题列表
            root_title: 根节点标题
            sheet_title: 工作表标题

        Returns:
            dict: 完整的 XMind JSON 结构
        """
        return {
            'id': self._generate_id(),
            'title': root_title,
            'topic': {
                'id': self._generate_id(),
                'link': None,
                'title': sheet_title,
                'note': None,
                'label': None,
                'comment': None,
                'markers': [],
                'topics': test_point_topics
            }
        }

    @staticmethod
    def _generate_id():
        """生成唯一 ID"""
        return str(uuid.uuid4())

    def save_xmind_json(self, data, output_file):
        """
        将 XMind 格式的 JSON 数据保存到文件

        Args:
            data: JSON 数据
            output_file: 输出文件路径

        Returns:
            bool: 保存是否成功
        """
        try:
            self._ensure_directory_exists(output_file)

            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False)

            self.logger.info(f"XMind 格式 JSON 已成功保存到: {output_file}")
            return True
        except Exception as e:
            self.logger.error(f"保存 JSON 文件时发生错误: {str(e)}", exc_info=True)
            return False

    def convert_and_export_to_xmind(self, input_data, output_xmind_file,
                                    root_title=DEFAULT_ROOT_TITLE,
                                    sheet_title=DEFAULT_SHEET_TITLE,
                                    template_file=None):

        # 步骤1: 转换为 XMind 标准 JSON 格式
        xmind_json = self.convert_test_points_to_xmind_format(
            input_data,
            root_title,
            sheet_title
        )

        # 步骤2: 导出为 XMind 文件
        success = self.json_to_xmind(xmind_json, output_xmind_file, template_file)

        if success:
            self.logger.info(f"XMind 文件已成功生成: {output_xmind_file}")

        return success

    @staticmethod
    def _ensure_directory_exists(file_path):
        """
        确保文件所在目录存在

        Args:
            file_path: 文件路径
        """
        file_dir = os.path.dirname(file_path)
        if file_dir and not os.path.exists(file_dir):
            os.makedirs(file_dir, exist_ok=True)


#获取测试用例转化为xmind格式文件
class TestcaseJsonToXmind:
    """将测试用例JSON数据转换为XMind格式的处理器"""

    DEFAULT_ROOT_TITLE = "测试用例"
    DEFAULT_SHEET_TITLE = "功能测试"
    DEFAULT_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'template.xmind')
    BLANK_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'blank_template.xmind')
    FALLBACK_TEMPLATE_PATH = r"C:\Users\admin\Desktop\test.xmind"

    def __init__(self, template_file=None):
        """
        初始化转换器

        Args:
            template_file: XMind 模板文件路径，默认使用内置模板
        """
        self.logger = LogManager().get_logger() if hasattr(LogManager(), 'get_logger') else logger
        self._ensure_blank_template_exists()
        self.template_file = template_file or self.BLANK_TEMPLATE_PATH

    def _ensure_blank_template_exists(self):
        """确保空白模板文件存在，如果不存在则自动创建"""
        if os.path.exists(self.BLANK_TEMPLATE_PATH):
            return

        try:
            import zipfile

            with zipfile.ZipFile(self.BLANK_TEMPLATE_PATH, 'w', zipfile.ZIP_DEFLATED) as zf:
                content_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="no"?>
<xmap-content xmlns="urn:xmind:xmap:xmlns:content:2.0" xmlns:fo="http://www.w3.org/1999/XSL/Format" xmlns:svg="http://www.w3.org/2000/svg" xmlns:xhtml="http://www.w3.org/1999/xhtml" xmlns:xlink="http://www.w3.org/1999/xlink" modified-by="Lingma" timestamp="1234567890">
<sheet id="sheet1" modified-by="Lingma" timestamp="1234567890">
<topic id="root1" modified-by="Lingma" structure-class="org.xmind.ui.logic.right" timestamp="1234567890">
<title>Root</title>
</topic>
</sheet>
</xmap-content>'''
                zf.writestr('content.xml', content_xml.encode('utf-8'))

                meta_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="no"?>
<meta xmlns="urn:xmind:xmap:xmlns:meta:2.0" version="2.0">
<Creator>Lingma AI</Creator>
<Created>2024-01-01T00:00:00.000+0800</Created>
</meta>'''
                zf.writestr('meta.xml', meta_xml.encode('utf-8'))

            self.logger.info(f"✓ 已自动创建空白模板：{self.BLANK_TEMPLATE_PATH}")
        except Exception as e:
            self.logger.error(f"创建空白模板失败：{str(e)}")
            raise

    def _get_template_file(self, template_file=None):
        """获取有效的模板文件路径"""
        if template_file and os.path.exists(template_file):
            return template_file
        elif os.path.exists(self.BLANK_TEMPLATE_PATH):
            return self.BLANK_TEMPLATE_PATH
        elif os.path.exists(self.DEFAULT_TEMPLATE_PATH):
            return self.DEFAULT_TEMPLATE_PATH
        elif os.path.exists(self.FALLBACK_TEMPLATE_PATH):
            return self.FALLBACK_TEMPLATE_PATH
        else:
            raise FileNotFoundError(
                f"未找到 XMind 模板文件。\n"
                f"请确保以下路径之一存在：\n"
                f"1. {self.BLANK_TEMPLATE_PATH}（自动生成）\n"
                f"2. {self.DEFAULT_TEMPLATE_PATH}\n"
                f"3. {self.FALLBACK_TEMPLATE_PATH}"
            )

    def json_to_xmind(self, json_data, output_file, template_file=None):
        """
        将测试用例JSON数据转换为XMind文件

        Args:
            json_data: 测试用例JSON数据（字典或列表）
            output_file: 输出的XMind文件路径
            template_file: 可选的模板文件路径

        Returns:
            bool: 转换是否成功
        """
        try:
            valid_template = self._get_template_file(template_file)

            workbook = xmind.load(valid_template)
            sheet = workbook.getPrimarySheet()

            if not sheet:
                self.logger.error("模板文件中没有工作表")
                return False

            sheet_title = json_data.get('sheet_title', self.DEFAULT_SHEET_TITLE) if isinstance(json_data, dict) else self.DEFAULT_SHEET_TITLE
            sheet.setTitle(sheet_title)

            root_topic = sheet.getRootTopic()

            self._build_testcase_structure(root_topic, json_data)

            self._ensure_directory_exists(output_file)

            xmind.save(workbook, output_file)
            self.logger.info(f"XMind 文件已成功生成：{output_file}")
            return True

        except FileNotFoundError as e:
            self.logger.error(f"模板文件未找到：{str(e)}")
            return False
        except Exception as e:
            self.logger.error(f"转换 XMind 文件时发生错误：{str(e)}", exc_info=True)
            return False


    def _build_testcase_structure(self, root_topic, json_data):
        """
        构建测试用例的XMind结构

        Args:
            root_topic: XMind根主题对象
            json_data: 测试用例JSON数据
        """
        if isinstance(json_data, dict):
            test_cases = json_data.get('测试用例', [])
            if not test_cases:
                test_cases = json_data.get('testcases', [])

            if isinstance(test_cases, list):
                # 按模块分组
                modules_dict = {}
                for testcase in test_cases:
                    if isinstance(testcase, dict):
                        module_name = testcase.get('测试模块', '未分类模块')
                        if module_name not in modules_dict:
                            modules_dict[module_name] = []
                        modules_dict[module_name].append(testcase)

                # 为每个模块创建节点
                for module_name, cases in modules_dict.items():
                    module_topic = root_topic.addSubTopic()
                    module_topic.setTitle(module_name)

                    # 为该模块下的每个用例创建子节点
                    for testcase in cases:
                        self._add_testcase_hierarchy(module_topic, testcase)
            else:
                self.logger.warning("测试用例数据格式不正确，应为列表")

        elif isinstance(json_data, list):
            # 按模块分组
            modules_dict = {}
            for testcase in json_data:
                if isinstance(testcase, dict):
                    module_name = testcase.get('测试模块', '未分类模块')
                    if module_name not in modules_dict:
                        modules_dict[module_name] = []
                    modules_dict[module_name].append(testcase)

            # 为每个模块创建节点
            for module_name, cases in modules_dict.items():
                module_topic = root_topic.addSubTopic()
                module_topic.setTitle(module_name)

                # 为该模块下的每个用例创建子节点
                for testcase in cases:
                    self._add_testcase_hierarchy(module_topic, testcase)


    def _add_testcase_hierarchy(self, parent_topic, testcase):
        """
        添加测试用例的完整层级结构（模块 -> 标题 -> 前置条件 -> 操作步骤 -> 预期结果）

        Args:
            parent_topic: 父主题对象（模块节点）
            testcase: 测试用例数据
        """
        # 第1层：用例标题作为模块的子节点
        title = testcase.get('标题', '未命名用例')
        priority = testcase.get('用例等级', '')

        title_topic = parent_topic.addSubTopic()
        title_topic.setTitle(title)

        # 添加优先级图标（使用 XMind 原生图标）
        if priority:
            # 将 P0, P1, P2, P3 等映射到 XMind 的优先级图标
            priority_map = {
                'P0': MarkerId.priority1,  # 优先级1（红色旗帜）
                'P1': MarkerId.priority2,  # 优先级2（橙色旗帜）
                'P2': MarkerId.priority3,  # 优先级3（黄色旗帜）
                'P3': MarkerId.priority4,  # 优先级4（绿色旗帜）
                'P4': MarkerId.priority5,  # 优先级5（蓝色旗帜）
                'P5': MarkerId.priority6,  # 优先级6（紫色旗帜）
                'P6': MarkerId.priority7,  # 优先级7（灰色旗帜）
                'P7': MarkerId.priority8,  # 优先级8（白色旗帜）
                'P8': MarkerId.priority9,  # 优先级9
            }

            marker_id = priority_map.get(priority)
            if marker_id:
                try:
                    title_topic.addMarker(marker_id)
                except Exception as e:
                    self.logger.warning(f"添加优先级图标失败: {str(e)}")

        # 第2层：前置条件作为标题的子节点
        precondition = testcase.get('前置条件', '')
        if precondition:
            precondition_topic = title_topic.addSubTopic()

            # 如果前置条件是列表，在当前节点用编号展示
            if isinstance(precondition, list):
                precondition_items = [f"{i+1}. {str(p)}" for i, p in enumerate(precondition)]
                precondition_topic.setTitle("\n".join(precondition_items))
            else:
                precondition_topic.setTitle(f"{precondition}")

            # 第3层：操作步骤作为前置条件的子节点
            steps = testcase.get('操作步骤', [])
            if steps:
                steps_topic = precondition_topic.addSubTopic()

                # 如果操作步骤是列表，在当前节点用编号展示
                if isinstance(steps, list):
                    steps_items = [f"{str(step)}" for i, step in enumerate(steps)]
                    steps_topic.setTitle("\n".join(steps_items))
                else:
                    steps_topic.setTitle(f"{steps}")

                # 第4层：预期结果作为操作步骤的子节点
                expected_result = testcase.get('预期结果', '')
                if expected_result:
                    result_topic = steps_topic.addSubTopic()

                    # 如果预期结果是列表，在当前节点用编号展示
                    if isinstance(expected_result, list):
                        result_items = [f"{str(r)}" for i, r in enumerate(expected_result)]
                        result_topic.setTitle("\n".join(result_items))
                    else:
                        result_topic.setTitle(f"{expected_result}")
            else:
                # 如果没有操作步骤，但有预期结果，预期结果作为前置条件的子节点
                expected_result = testcase.get('预期结果', '')
                if expected_result:
                    result_topic = precondition_topic.addSubTopic()

                    if isinstance(expected_result, list):
                        result_items = [f"{i+1}. {str(r)}" for i, r in enumerate(expected_result)]
                        result_topic.setTitle("预期结果：\n" + "\n".join(result_items))
                    else:
                        result_topic.setTitle(f"{expected_result}")
        else:
            # 如果没有前置条件，操作步骤作为标题的子节点
            steps = testcase.get('操作步骤', [])
            if steps:
                steps_topic = title_topic.addSubTopic()

                if isinstance(steps, list):
                    steps_items = [f"{str(step)}" for i, step in enumerate(steps)]
                    steps_topic.setTitle("\n".join(steps_items))
                else:
                    steps_topic.setTitle(f"{steps}")

                # 预期结果作为操作步骤的子节点
                expected_result = testcase.get('预期结果', '')
                if expected_result:
                    result_topic = steps_topic.addSubTopic()

                    if isinstance(expected_result, list):
                        result_items = [f"{str(r)}" for i, r in enumerate(expected_result)]
                        result_topic.setTitle("\n".join(result_items))
                    else:
                        result_topic.setTitle(f"{expected_result}")
            else:
                # 如果都没有，直接添加预期结果
                expected_result = testcase.get('预期结果', '')
                if expected_result:
                    result_topic = title_topic.addSubTopic()

                    if isinstance(expected_result, list):
                        result_items = [f"{str(r)}" for i, r in enumerate(expected_result)]
                        result_topic.setTitle("\n".join(result_items))
                    else:
                        result_topic.setTitle(f"{expected_result}")

    @staticmethod
    def extract_priority_from_title(title):
        """
        从标题中提取优先级标记

        Args:
            title: 标题字符串，如 "[P0] 测试模块1"

        Returns:
            tuple: (优先级, 清理后的标题)
                   例如: ("P0", "测试模块1")
        """
        # 匹配 [P0], [P1], [P2], [P3] 等格式
        pattern = r'^\[(P\d+)\]\s*(.+)$'
        match = re.match(pattern, title)

        if match:
            priority = match.group(1)
            clean_title = match.group(2)
            return priority, clean_title
        else:
            return None, title

    @staticmethod
    def add_priority_to_title(priority, title):
        """
        给标题添加优先级标记（已废弃，现在使用图标方式）

        Args:
            priority: 优先级，如 "P0", "P1"
            title: 原始标题

        Returns:
            str: 带优先级的标题（仅用于兼容旧代码）
        """
        if priority:
            return f"[{priority}] {title}"
        return title


    def convert_and_export_to_xmind(self, input_data, output_xmind_file,
                                    root_title=DEFAULT_ROOT_TITLE,
                                    sheet_title=DEFAULT_SHEET_TITLE,
                                    template_file=None):
        """
        一键转换并导出测试用例为XMind文件

        Args:
            input_data: 测试用例JSON数据
            output_xmind_file: 输出的XMind文件路径
            root_title: 根节点标题
            sheet_title: 工作表标题
            template_file: 可选的模板文件路径

        Returns:
            bool: 转换是否成功
        """
        try:
            if isinstance(input_data, dict):
                input_data['sheet_title'] = sheet_title

            success = self.json_to_xmind(input_data, output_xmind_file, template_file)

            if success:
                self.logger.info(f"XMind 文件已成功生成: {output_xmind_file}")

            return success

        except Exception as e:
            self.logger.error(f"转换过程中发生错误: {str(e)}", exc_info=True)
            return False

    @staticmethod
    def _ensure_directory_exists(file_path):
        """
        确保文件所在目录存在

        Args:
            file_path: 文件路径
        """
        file_dir = os.path.dirname(file_path)
        if file_dir and not os.path.exists(file_dir):
            os.makedirs(file_dir, exist_ok=True)


#测试用例的xmind格式文件转为xls/csv/xlsx格式文件
# class TestcaseXmindToXlsx:



#AI给出json，转化为execl格式文件






if __name__ == '__main__':
    # dic_to_xlsx = DicToXlsx()
    # dic_to_xlsx.table_data_processing()
    # dc = MxindDataProcessor()
    # res = dc.xmind_to_json()
    # DD = WriteInfo()
    # DD.write_json_to_file(data=res, file=r"D:\AIGeneration\testcase\demo.json")
    #
    # with open(r"D:\AIGeneration\testcase\demo.json", 'r', encoding='utf-8') as f:
    #     json_data = json.load(f)
    # # # 转换为 XMind并导出
    # output_xmind = r"C:\Users\admin\Desktop\demo_output.xmind"
    # json_to_xmind(json_data, output_xmind)
    # res = XmindPointJson()
    # res.extract_test_points_data()
    # converter = JsonToXmind()
    # # 准备测试点数据
    # filepath =r"D:\AIGeneration\testcase\测试点.json"
    # get_json = fileProcessor()
    # test_data = get_json.find_and_read_file(filepath, type="json")
    # # 一键转换并导出
    # success = converter.convert_and_export_to_xmind(
    #     input_data=test_data,
    #     output_xmind_file=r"D:\AIGeneration\testcase\output.xmind",
    #     root_title="测试大纲",
    #     sheet_title="功能测试用例"
    # )
    # 测试用例JSON转XMind示例
    print("\n=== 测试用例JSON转XMind ===")
    testcase_converter = TestcaseJsonToXmind()

    # 读取测试用例JSON文件
    testcase_json_path = r"D:\AIGeneration\testcase\测试用例.json"
    get_json = fileProcessor()
    testcase_data = get_json.find_and_read_file(testcase_json_path, type="json")

    # 转换并导出
    output_xmind = r"D:\AIGeneration\testcase\测试用例.xmind"
    success = testcase_converter.convert_and_export_to_xmind(
        input_data=testcase_data,
        output_xmind_file=output_xmind,
        root_title="测试用例",
        sheet_title="功能测试用例"
    )

    if success:
        print(f"✓ 测试用例XMind文件已生成：{output_xmind}")
    else:
        print("✗ 转换失败")