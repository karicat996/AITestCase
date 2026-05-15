# mxind数据转换器
import xmind
import json
import re
import uuid
from xmind.core.markerref import MarkerId
from utils.logs import LogManager
from common.textRecognition import *
from loguru import logger
from common.fileProcessor import *
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
import os

fp = fileProcessor()
OUTPUT_JSON_PATH = fp.find_and_read_file("config/systemConfig.yaml", type="yaml").get("OUTPUT_JSON_PATH")
DEFAULT_TEMPLATE_PATH = fp.find_and_read_file("config/systemConfig.yaml", type="yaml").get("DEFAULT_TEMPLATE_PATH")
TEMPLATE_PATH = fp.find_and_read_file("config/systemConfig.yaml", type="yaml").get("TEMPLATE_PATH")
TEST_XMIND_PATH = fp.find_and_read_file("config/systemConfig.yaml", type="yaml").get("TEST_XMIND_PATH")
TEMPLATE_XMIND_PATH = fp.find_and_read_file("config/systemConfig.yaml", type="yaml").get("TEMPLATE_XMIND_PATH")
TESTCASE_JSON_PATH = fp.find_and_read_file("config/systemConfig.yaml", type="yaml").get("TESTCASE_JSON_PATH")
OUTPUT_XLSX_PATH = fp.find_and_read_file("config/systemConfig.yaml", type="yaml").get("OUTPUT_XLSX_PATH")
CONVERTED_TESTCASES_JSON_PATH = fp.find_and_read_file("config/systemConfig.yaml", type="yaml").get("CONVERTED_TESTCASES_JSON_PATH")
IMG_PATH = fp.find_and_read_file("config/systemConfig.yaml", type="yaml").get("IMG_PATH")
TEST_POINT_XMIND_FILE = fp.find_and_read_file("config/systemConfig.yaml", type="yaml").get("TEST_POINT_XMIND_FILE")
LogManager(log_dir=r"D:\AIGeneration\utils\logs")


def _build_xmind_topic(xmind_topic, data):
    """
    递归构建 XMind 主题

    Args:
        xmind_topic: XMind 主题对象
        data: 数据字典或字符串
    """
    if isinstance(data, str):
        xmind_topic.setTitle(data)
    elif isinstance(data, dict):
        # 设置当前主题标题
        title = data.get('title', data.get('name', ''))
        if title:
            xmind_topic.setTitle(title)

        # 处理子主题
        topics = data.get('topics', data.get('children', []))
        if topics and isinstance(topics, list):
            for child_data in topics:
                child_topic = xmind_topic.addSubTopic()
                _build_xmind_topic(child_topic, child_data)

        # 处理备注信息（如果有）
        notes = data.get('notes', data.get('comment', ''))
        if notes:
            xmind_topic.setNote(notes)

        # 处理测试用例的特殊字段
        if 'steps' in data or 'result' in data:
            note_content = ""
            if 'steps' in data and data['steps']:
                note_content += f"步骤：{data['steps']}\n"
            if 'result' in data and data['result']:
                note_content += f"预期结果：{data['result']}\n"
            if note_content:
                existing_note = xmind_topic.getNote()
                if existing_note:
                    xmind_topic.setNote(f"{existing_note}\n{note_content}")
                else:
                    xmind_topic.setNote(note_content)
    elif isinstance(data, list):
        # 如果是列表，第一个元素作为标题，其余作为子主题
        if len(data) > 0:
            xmind_topic.setTitle(str(data[0]))
            for item in data[1:]:
                child_topic = xmind_topic.addSubTopic()
                _build_xmind_topic(child_topic, item)

#  xmind数据整理
class MxindDataProcessor:
    def __init__(self):
        self.xmind_file = TEST_XMIND_PATH
        self.logging = LogManager()
        self.case_dict = {}
    def xmind_to_json(self):
        workbook = xmind.load(self.xmind_file)
        sheet = workbook.getData()
        logger.debug("调试信息")
        if sheet:
            logger.debug("调试信息")
            data = sheet[0]
            return data
            print(data)
    def write_to_json(self,output_json_path):
        res = self.xmind_to_json()
        # 获取到数据后写入xmind.json文件
        if res:
            with open(output_json_path, 'w', encoding='utf-8') as f:
                json.dump(res, f, ensure_ascii=False, indent=2)
            print(f"✓ XMind数据已保存到: {output_json_path}")
        else:
            print("✗ 未能读取到XMind数据")

#测试用例数据提取
class AdvancedTestCaseExtractor:

    def __init__(self, max_depth=None):
        self.max_depth = max_depth
        self.results = []

    def extract(self, json_obj):
        """
        提取测试用例数据
        """
        self.results = []
        root_topic = json_obj.get('topic', {})

        if root_topic:
            # 从根节点开始遍历，传递路径信息
            self._traverse_with_path(root_topic, [], 1)

        return self.results

    def _traverse_with_path(self, node, path, level):
        """
        带路径信息的遍历方法 - 基于层级位置判断
        """
        if self.max_depth is not None and level > self.max_depth:
            return

        # 更新路径
        current_path = path + [node]

        # 基于层级位置判断节点类型
        # 假设结构：第1层为产品名，第2层为模块名，第3层为测试用例
        if level == 4:  # 测试用例节点位于第3层
            # 创建基础测试用例信息
            test_case_info = self._create_test_case_dict_by_position(node, current_path, level)

            # 获取按步骤拆分后的结果
            split_entries = self._extract_steps_and_results_by_position(node, test_case_info)

            if split_entries:
                # 添加拆分后的条目
                self.results.extend(split_entries)
            else:
                # 如果没有子节点，添加原始条目
                self.results.append(test_case_info)

        # 递归处理子节点
        for child in node.get('topics', []):
            self._traverse_with_path(child, current_path, level + 1)

    def _create_test_case_dict_by_position(self, test_case_node, path, level):
        """
        根据节点在树中的位置创建测试用例字典
        """
        subject = ''
        model = ''

        #  这里处理节点的属性，例如主题、模型等
        if len(path) >= 1:
            subject = path[0].get('title', '')
        if len(path) >= 2:
            model = path[1].get('title', '')
        if len(path) >= 3:
            title = path[2].get('title', '')
        if len(path) >= 4:
            header = path[3].get('title', '')
        return {
            'Subject': subject,
            'model': model,
            'title': title,
            'header': header,
            'steps': '',
            'result': ''
        }

    def _extract_steps_and_results_by_position(self, test_case_node, test_case_info):
        """
        基于位置提取测试步骤和预期结果
        假设第一层子节点是步骤，第二层子节点是对应的结果
        """
        new_entries = []

        # 获取子节点（假设为测试步骤）
        child_nodes = test_case_node.get('topics', [])

        for step_node in child_nodes:
            # 获取步骤节点的子节点（假设为预期结果）
            result_nodes = step_node.get('topics', [])

            if result_nodes:
                # 对每个结果节点创建一个条目
                for result_node in result_nodes:
                    new_entry = {
                        'Subject': test_case_info['Subject'],
                        'model': test_case_info['model'],
                        'title': test_case_info['title'],
                        'header': test_case_info['header'],
                        'steps': step_node.get('title', ''),
                        'result': result_node.get('title', '')
                    }
                    new_entries.append(new_entry)
            else:
                # 如果步骤节点没有结果子节点
                new_entry = {
                    'Subject': test_case_info['Subject'],
                    'model': test_case_info['model'],
                    'title': test_case_info['title'],
                    'header': test_case_info['header'],
                    'steps': step_node.get('title', ''),
                    'result': ''
                }
                new_entries.append(new_entry)

        return new_entries

    def get_statistics(self):
        """
        获取提取结果的统计信息

        Returns:
            dict: 统计信息
        """
        if not self.results:
            return {'total_cases': 0}

        total_steps = len(self.results)  # 每个条目代表一个步骤-结果对
        total_results = len(self.results)

        data = {
            'total_cases': len(self.results),
            'total_steps': total_steps,
            'total_results': total_results
        }

        print(data)

#字典转化成xlsx
class DicToXlsx:
    def __init__(self):
        self.xlsx_file = r'C:\Users\admin\Desktop\demo.xlsx'
        file_dir = os.path.dirname(self.xlsx_file)
        if file_dir and not os.path.exists(file_dir):  # 只有当目录不存在时才创建
            os.makedirs(file_dir, exist_ok=True)
        json_data = MxindDataProcessor().xmind_to_json()
        extractor = AdvancedTestCaseExtractor(max_depth=10)
        self.data = extractor.extract(json_data)
        statistics = extractor.get_statistics()
        print(f"提取到的数据: {len(self.data)} 条")

    def table_data_processing(self):
        try:
            # 如果文件不存在，则创建新的工作簿并设置表头
            if not os.path.exists(self.xlsx_file):
                wb = Workbook()
                sheet = wb.active
                # 定义中文表头
                headers_chinese = ['主题', '模块', '标题', '前置条件', '步骤', '预期结果']
                for col_num, header in enumerate(headers_chinese, 1):
                    sheet.cell(row=1, column=col_num, value=header)
                # 设置表头行高
                sheet.row_dimensions[1].height = 25
                wb.save(self.xlsx_file)
                print(f"已创建新文件并写入表头: {self.xlsx_file}")

            # 加载现有的工作簿
            wb = load_workbook(self.xlsx_file)
            sheet = wb.active

            # 检查表头是否已经存在，如果不存在则添加
            first_cell = sheet.cell(row=1, column=1).value
            headers_chinese = ['主题', '模块', '标题', '前置条件', '步骤', '预期结果']

            # 检查第一列是否包含期望的表头值之一
            if not first_cell or first_cell not in headers_chinese:
                # 重新写入表头
                for col_num, header in enumerate(headers_chinese, 1):
                    sheet.cell(row=1, column=col_num, value=header)
                print("已写入表头到第一行")

            # 获取现有数据的最后一行
            last_row = sheet.max_row
            # 定义英文键名用于从数据字典中获取值
            headers_english = ['Subject', 'model', 'title', 'header', 'steps', 'result']

            # 遍历数据数组，逐行写入
            for row_idx, data_dict in enumerate(self.data, start=last_row + 1):
                for col_idx, header in enumerate(headers_english, start=1):
                    # 从字典中获取对应键的值，如果键不存在则写入空字符串
                    value = data_dict.get(header, '')
                    sheet.cell(row=row_idx, column=col_idx, value=value)

            # 保存工作簿
            wb.save(self.xlsx_file)
            print(f"数据已成功写入: {self.xlsx_file}")

        except PermissionError:
            print(f"权限错误：无法访问文件 {self.xlsx_file}，请确保文件未被其他程序打开")
        except Exception as e:
            print(f"写入Excel文件时发生错误: {str(e)}")

# 测试点的xmind 转化 json
class XmindPointJson:
    """
    XMind测试点数据转换为JSON格式的工具类
    
    功能包括：
    - 读取XMind文件数据
    - 过滤和提取测试点数据
    - 将处理后的数据保存为JSON文件
    """

    def __init__(self, xmind_file=None):
        """
        初始化XMind测试点转换器
        
        Args:
            xmind_file: XMind文件路径，默认使用配置文件中的TEST_POINT_XMIND_FILE
        """
        self.xmind_file = xmind_file or TEST_POINT_XMIND_FILE


    def read_xmind_data(self):
        """
        读取XMind文件数据
        
        Returns:
            dict: XMind文件的JSON数据，如果读取失败返回None
        """
        try:
            if not os.path.exists(self.xmind_file):
                logger.error(f"XMind文件不存在: {self.xmind_file}")
                return None
            
            workbook = xmind.load(self.xmind_file)
            sheet = workbook.getData()
            
            if not sheet:
                logger.warning("模板文件中没有工作表")
                return None
            
            json_data = sheet[0]
            logger.info(f"成功读取XMind数据，包含 {len(json_data.get('topic', {}).get('topics', []))} 个主题")
            return json_data
            
        except Exception as e:
            logger.error(f"读取XMind文件失败: {str(e)}", exc_info=True)
            return None

    def _extract_test_points_data(self, data=None):
        """
        从XMind数据中提取测试点和其子项，生成符合测试点.json格式的深层嵌套JSON
        Args:
            data: 包含原始数据的字典，如果为None则自动读取XMind文件
            
        Returns:
            dict: 符合测试点.json格式的深层嵌套JSON数据
        """
        # 如果没有提供数据，则自动读取XMind文件
        if data is None:
            data = self.read_xmind_data()
            
        if not data:
            logger.error("无法获取有效的XMind数据")
            return {}
        
        try:
            # 创建结果字典
            result = {}
            
            # 遍历主题下的topics获取顶层产品/需求名称
            topics = data.get('topic', {}).get('topics', [])
            
            if not topics:
                logger.error("XMind数据中没有找到任何主题节点，请检查XMind文件结构")
                return {}
            
            for topic in topics:
                # 第一层：产品/需求名称（如"电商下单界面功能"）
                product_name = topic.get('title')
                
                if not product_name:
                    logger.warning("发现无标题的主题节点，跳过")
                    continue
                
                logger.debug(f"正在处理产品: {product_name}")
                
                # 获取该产品的所有模块
                module_topics = topic.get('topics', [])
                
                if not module_topics:
                    logger.warning(f"产品 '{product_name}' 下没有模块，跳过")
                    continue
                
                # 第二层：模块列表
                modules_list = []
                
                for module_topic in module_topics:
                    # 第二层：模块名称（如"商品选择模块"）
                    module_name = module_topic.get('title')
                    
                    if not module_name:
                        logger.warning(f"在产品 '{product_name}' 中发现无标题的模块，跳过")
                        continue
                    
                    logger.debug(f"正在处理模块: {module_name}")
                    
                    # 获取该模块下的所有测试点
                    test_point_topics = module_topic.get('topics', [])
                    
                    if not test_point_topics:
                        logger.warning(f"模块 '{module_name}' 下没有测试点，跳过")
                        continue
                    
                    # 第三层：测试点列表
                    test_points_list = []
                    
                    for test_point_topic in test_point_topics:
                        # 第三层：测试点名称（如"商品添加与数量修改"）
                        test_point_name = test_point_topic.get('title')
                        
                        if not test_point_name:
                            logger.warning(f"在模块 '{module_name}' 中发现无标题的测试点，跳过")
                            continue
                        
                        # 获取该测试点下的所有场景
                        scenario_topics = test_point_topic.get('topics', [])
                        
                        if not scenario_topics:
                            logger.warning(f"测试点 '{test_point_name}' 下没有场景，跳过")
                            continue
                        
                        # 第四层：场景描述列表
                        scenarios_list = []
                        
                        for scenario_topic in scenario_topics:
                            scenario_title = scenario_topic.get('title')
                            scenario_note = scenario_topic.get('note', '') or scenario_topic.get('comment', '')
                            
                            if not scenario_title:
                                logger.warning(f"在测试点 '{test_point_name}' 中发现无标题的场景，跳过")
                                continue
                            
                            # 构建场景字典：{场景标题: 场景描述}
                            scenario_dict = {scenario_title: scenario_note if scenario_note else ""}
                            scenarios_list.append(scenario_dict)
                        
                        if scenarios_list:
                            # 将测试点和其场景添加到测试点列表
                            test_points_list.append({test_point_name: scenarios_list})
                    
                    if test_points_list:
                        # 将模块和其测试点添加到模块列表
                        modules_list.append({module_name: test_points_list})
                
                if modules_list:
                    # 将产品和其模块添加到结果中
                    result[product_name] = modules_list
            
            if not result:
                logger.error("未能提取到任何有效的测试点数据，请检查XMind文件结构是否符合要求")
                return {}
            
            logger.info(f"成功提取 {len(result)} 个产品，共 {sum(len(modules) for modules in result.values())} 个模块")
            return result
            
        except Exception as e:
            logger.error(f"提取测试点数据时发生错误: {str(e)}", exc_info=True)
            return {}

    def save_to_json(self, data, output_file=None):
        """
        将处理后的数据保存为JSON文件
        
        Args:
            data: 要保存的数据字典
            output_file: 输出文件路径，如果为None则使用默认路径
            
        Returns:
            bool: 保存是否成功
        """
        if not data:
            logger.warning("没有数据需要保存")
            return False
        
        # 确定输出文件路径
        if output_file is None:
            # 使用默认输出路径
            output_dir = os.path.dirname(TEST_POINT_XMIND_FILE)
            output_file = os.path.join(output_dir, 'test_points_output.json')
        
        try:
            # 确保输出目录存在
            output_dir = os.path.dirname(output_file)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
                logger.info(f"创建输出目录: {output_dir}")
            
            # 写入JSON文件
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            logger.info(f"测试点数据已保存到: {output_file}")
            print(f"✓ 测试点数据已保存到: {output_file}")
            return True
            
        except Exception as e:
            logger.error(f"保存JSON文件失败: {str(e)}", exc_info=True)
            print(f"✗ 保存失败: {str(e)}")
            return False

    def process_and_save(self, output_file=None):
        """
        完整流程：读取XMind -> 提取测试点 -> 保存为JSON
        
        Args:
            output_file: 输出JSON文件路径，可选
            
        Returns:
            dict: 提取的测试点数据，如果失败返回空字典
        """
        logger.info("开始处理XMind测试点数据...")
        
        # 步骤1: 读取XMind数据
        xmind_data = self.read_xmind_data()
        if not xmind_data:
            logger.error("读取XMind数据失败")
            return {}
        
        # 步骤2: 提取测试点数据
        test_points_data = self._extract_test_points_data(xmind_data)
        if not test_points_data:
            logger.error("提取测试点数据失败")
            return {}
        
        # 步骤3: 保存为JSON文件
        success = self.save_to_json(test_points_data, output_file)
        
        if success:
            logger.info(f"处理完成！共提取 {len(test_points_data)} 个测试点")
            print(f"✓ 处理完成！共提取 {len(test_points_data)} 个测试点")
        else:
            logger.error("保存JSON文件失败")
            print("✗ 处理失败")
        
        return test_points_data

#将AI给的json数据转化为xmind
class TestcasePointJsonToXmind:
    """JSON 数据转换为 XMind 格式的处理器"""

    DEFAULT_ROOT_TITLE = "逻辑图"
    DEFAULT_SHEET_TITLE = "测试点"
    # DEFAULT_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'template.xmind')
    # BLANK_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'blank_template.xmind')
    DEFAULT_TEMPLATE_PATH = DEFAULT_TEMPLATE_PATH
    BLANK_TEMPLATE_PATH = TEMPLATE_PATH
    FALLBACK_TEMPLATE_PATH = TEMPLATE_XMIND_PATH
    # 标记是否已经警告过模板残留问题
    _template_warning_shown = False
    def __init__(self, template_file=None):
        """
        初始化转换器

        Args:
            template_file: XMind 模板文件路径，默认使用内置模板
        """
        self._ensure_blank_template_exists()
        self.template_file = template_file or self.BLANK_TEMPLATE_PATH


    def _ensure_blank_template_exists(self):
        """确保空白模板文件存在，如果不存在则自动创建"""
        if os.path.exists(self.BLANK_TEMPLATE_PATH):
            return  # 已存在，无需创建

        try:
            import zipfile

            # 创建一个真正的空白 XMind 文件（XMind 8 格式）
            with zipfile.ZipFile(self.BLANK_TEMPLATE_PATH, 'w', zipfile.ZIP_DEFLATED) as zf:
                # content.xml - 只包含一个空的根主题，没有任何子主题
                content_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="no"?>
    <xmap-content xmlns="urn:xmind:xmap:xmlns:content:2.0" xmlns:fo="http://www.w3.org/1999/XSL/Format" xmlns:svg="http://www.w3.org/2000/svg" xmlns:xhtml="http://www.w3.org/1999/xhtml" xmlns:xlink="http://www.w3.org/1999/xlink" modified-by="Lingma" timestamp="1234567890">
    <sheet id="sheet1" modified-by="Lingma" timestamp="1234567890">
    <topic id="root1" modified-by="Lingma" structure-class="org.xmind.ui.logic.right" timestamp="1234567890">
    <title>Root</title>
    </topic>
    </sheet>
    </xmap-content>'''
                zf.writestr('content.xml', content_xml.encode('utf-8'))

                # meta.xml - 元数据
                meta_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="no"?>
    <meta xmlns="urn:xmind:xmap:xmlns:meta:2.0" version="2.0">
    <Creator>Lingma AI</Creator>
    <Created>2024-01-01T00:00:00.000+0800</Created>
    </meta>'''
                zf.writestr('meta.xml', meta_xml.encode('utf-8'))

            logger.info(f"✓ 已自动创建空白模板：{self.BLANK_TEMPLATE_PATH}")
        except Exception as e:
            logger.error(f"创建空白模板失败：{str(e)}")
            raise
    def _get_template_file(self, template_file=None):

        # 优先级：用户指定 > 空白模板 > 默认模板 > 备用模板
        if template_file and os.path.exists(template_file):
            return template_file
        elif os.path.exists(self.BLANK_TEMPLATE_PATH):
            return self.BLANK_TEMPLATE_PATH
        elif os.path.exists(self.DEFAULT_TEMPLATE_PATH):
            return self.DEFAULT_TEMPLATE_PATH
        elif os.path.exists(self.FALLBACK_TEMPLATE_PATH):
            return self.FALLBACK_TEMPLATE_PATH
        else:
            raise FileNotFoundError(
                f"未找到 XMind 模板文件。\n"
                f"请确保以下路径之一存在：\n"
                f"1. {self.BLANK_TEMPLATE_PATH}（自动生成）\n"
                f"2. {self.DEFAULT_TEMPLATE_PATH}\n"
                f"3. {self.FALLBACK_TEMPLATE_PATH}"
            )

    def json_to_xmind(self, json_data, output_file, template_file=None):

        try:
            # 获取有效的模板文件
            valid_template = self._get_template_file(template_file)

            # 加载模板工作簿
            workbook = xmind.load(valid_template)
            sheet = workbook.getPrimarySheet()

            if not sheet:
                self.logger.error("模板文件中没有工作表")
                return False

            # 设置工作表标题
            sheet_title = json_data.get('title', self.DEFAULT_SHEET_TITLE) if isinstance(json_data,
             dict) else self.DEFAULT_SHEET_TITLE
            sheet.setTitle(sheet_title)

            # 获取根主题
            root_topic = sheet.getRootTopic()


            self._build_root_content_clean(root_topic, json_data)

            # 确保输出目录存在
            self._ensure_directory_exists(output_file)

            # 保存 XMind 文件
            xmind.save(workbook, output_file)
            logger.info(f"XMind 文件已成功生成：{output_file}")
            return True

        except FileNotFoundError as e:
            logger.error(f"模板文件未找到：{str(e)}")
            return False
        except Exception as e:
            logger.error(f"转换 XMind 文件时发生错误：{str(e)}", exc_info=True)
            return False


    def _build_root_content_clean(self, root_topic, json_data):
        """

        Args:
            root_topic: XMind 根主题对象
            json_data: JSON 数据
        """
        if isinstance(json_data, dict):
            if 'topic' in json_data:
                topic_data = json_data['topic']
                # 先设置根主题标题
                if 'title' in topic_data:
                    root_topic.setTitle(topic_data['title'])
                # 然后递归构建子主题
                _build_xmind_topic(root_topic, topic_data)
            else:
                # 直接作为根主题处理
                if 'title' in json_data:
                    root_topic.setTitle(json_data['title'])
                _build_xmind_topic(root_topic, json_data)
        elif isinstance(json_data, list):
            # 如果是列表，每个元素作为一个子主题
            for item in json_data:
                if isinstance(item, dict):
                    child_topic = root_topic.addSubTopic()
                    _build_xmind_topic(child_topic, item)

    def convert_test_points_to_xmind_format(self, input_data, root_title=DEFAULT_ROOT_TITLE, sheet_title=DEFAULT_SHEET_TITLE):
        """
        将测试点数据转换为 XMind 标准 JSON 格式
        支持多种模板规则自动识别
        """
        category_topics = []
        
        # 获取当前数据适用的提取规则
        rule = self._get_template_rule(input_data)

        # 遍历顶层分类（如"LCD光固化3D打印机"）
        for category_name, category_data in input_data.items():
            # 使用选定的规则提取子主题
            sub_topics = self._extract_sub_topics_by_rule(category_data, rule)
            
            category_topic = {
                'id': self._generate_id(),
                'link': None,
                'title': category_name,
                'note': None,
                'label': None,
                'comment': None,
                'markers': [],
                'topics': sub_topics
            }
            category_topics.append(category_topic)

        return self._build_xmind_json_structure(
            category_topics,
            root_title,
            sheet_title
        )

    def _get_template_rule(self, data):
        """
        根据数据结构自动判断使用的模板规则
        """
        if not data or not isinstance(data, dict):
            return 'simple'
        
        # 检查第一层值的类型
        first_val = next(iter(data.values()))
        if isinstance(first_val, list) and first_val and isinstance(first_val[0], dict):
            # 检查第二层是否还是字典列表（深度嵌套结构，如测试点.json）
            second_val = first_val[0]
            if isinstance(second_val, dict):
                return 'deep_nested'
        return 'standard'

    def _extract_sub_topics_by_rule(self, data, rule):
        """
        根据规则提取子主题
        """
        if rule == 'deep_nested':
            return self._extract_deep_topics(data)
        else:
            return self._extract_sub_topics(data)

    def _extract_deep_topics(self, data_list):
        """
        专门处理深层嵌套的字典结构 (Module -> SubModule -> TestPoints)
        对应结构: [{"Z轴": [{"调平": [{"类型":...}]}]}]
        """
        topics = []
        if not isinstance(data_list, list):
            return topics

        for item in data_list:
            if not isinstance(item, dict):
                continue
            
            # 遍历当前层的 Key-Value (例如: "Z轴运动功能": [...])
            for key, value in item.items():
                parent_topic = self._create_sub_topic(key)
                
                # 递归处理下一层
                if isinstance(value, list):
                    children = self._extract_deep_topics(value)
                    if children:
                        parent_topic['topics'] = children
                elif isinstance(value, dict):
                    # 如果值是字典，继续深入
                    children = self._extract_deep_topics([value])
                    if children:
                        parent_topic['topics'] = children
                elif isinstance(value, str):
                    # 叶子节点
                    parent_topic['topics'] = [self._create_sub_topic(value)]
                
                topics.append(parent_topic)
        return topics

    def _create_category_topic(self, category_name, category_data):
        """
        创建分类主题（如"功能测试"）

        Args:
            category_name: 分类名称
            category_data: 分类下的数据（dict 或 list）

        Returns:
            dict: 分类主题数据
        """
        sub_topics = self._extract_sub_topics(category_data)

        return {
            'id': self._generate_id(),
            'link': None,
            'title': category_name,  # 使用实际分类名，而非"测试点1"
            'note': None,
            'label': None,
            'comment': None,
            'markers': [],
            'topics': sub_topics
        }

    def _extract_sub_topics(self, subcategories):
        """
        提取子主题列表

        Args:
            subcategories: 子类别数据（dict 或 list）

        Returns:
            list: 子主题列表
        """
        sub_topics = []

        if isinstance(subcategories, dict):
            for sub_key, sub_value in subcategories.items():
                # 如果值是列表，创建带子主题的主题节点
                if isinstance(sub_value, list):
                    parent_topic = self._create_sub_topic(sub_key)
                    # 为列表中的每个元素创建子主题
                    child_topics = []
                    for item in sub_value:
                        child_topics.append(self._create_sub_topic(str(item)))
                    parent_topic['topics'] = child_topics
                    sub_topics.append(parent_topic)
                else:
                    # 如果值是字符串或其他类型，直接创建主题
                    sub_topics.append(self._create_sub_topic(sub_key))
        elif isinstance(subcategories, list):
            for item in subcategories:
                sub_topics.append(self._create_sub_topic(str(item)))

        return sub_topics

    def _create_sub_topic(self, title):
        """
        创建单个子主题

        Args:
            title: 子主题标题

        Returns:
            dict: 子主题数据
        """
        return {
            'id': self._generate_id(),
            'link': None,
            'title': title,
            'note': None,
            'label': None,
            'comment': None,
            'markers': []
        }

    def _build_xmind_json_structure(self, test_point_topics, root_title, sheet_title):
        """
        构建完整的 XMind JSON 结构

        Args:
            test_point_topics: 测试点主题列表
            root_title: 根节点标题
            sheet_title: 工作表标题

        Returns:
            dict: 完整的 XMind JSON 结构
        """
        return {
            'id': self._generate_id(),
            'title': root_title,
            'topic': {
                'id': self._generate_id(),
                'link': None,
                'title': sheet_title,
                'note': None,
                'label': None,
                'comment': None,
                'markers': [],
                'topics': test_point_topics
            }
        }

    @staticmethod
    def _generate_id():
        """生成唯一 ID"""
        return str(uuid.uuid4())

    def save_xmind_json(self, data, output_file):
        """
        将 XMind 格式的 JSON 数据保存到文件

        Args:
            data: JSON 数据
            output_file: 输出文件路径

        Returns:
            bool: 保存是否成功
        """
        try:
            self._ensure_directory_exists(output_file)

            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False)

            logger.info(f"XMind 格式 JSON 已成功保存到: {output_file}")
            return True
        except Exception as e:
            logger.error(f"保存 JSON 文件时发生错误: {str(e)}", exc_info=True)
            return False

    def convert_and_export_to_xmind(self, input_data, output_xmind_file,
                                    root_title=DEFAULT_ROOT_TITLE,
                                    sheet_title=DEFAULT_SHEET_TITLE,
                                    template_file=None):

        # 步骤1: 转换为 XMind 标准 JSON 格式
        xmind_json = self.convert_test_points_to_xmind_format(
            input_data,
            root_title,
            sheet_title
        )

        # 步骤2: 导出为 XMind 文件
        success = self.json_to_xmind(xmind_json, output_xmind_file, template_file)

        if success:
            logger.info(f"XMind 文件已成功生成: {output_xmind_file}")

        return success

    @staticmethod
    def _ensure_directory_exists(file_path):
        """
        确保文件所在目录存在

        Args:
            file_path: 文件路径
        """
        file_dir = os.path.dirname(file_path)
        if file_dir and not os.path.exists(file_dir):
            os.makedirs(file_dir, exist_ok=True)

#获取测试用例转化为xmind格式文件
class TestcaseJsonToXmind:
    """将测试用例JSON数据转换为XMind格式的处理器"""

    DEFAULT_ROOT_TITLE = "测试用例"
    DEFAULT_SHEET_TITLE = "功能测试"
    DEFAULT_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'template.xmind')
    BLANK_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'blank_template.xmind')
    FALLBACK_TEMPLATE_PATH = TEMPLATE_XMIND_PATH

    def __init__(self, template_file=None):
        """
        初始化转换器

        Args:
            template_file: XMind 模板文件路径，默认使用内置模板
        """
        self._ensure_blank_template_exists()
        self.template_file = template_file or self.BLANK_TEMPLATE_PATH

    def _ensure_blank_template_exists(self):
        """确保空白模板文件存在，如果不存在则自动创建"""
        if os.path.exists(self.BLANK_TEMPLATE_PATH):
            return

        try:
            import zipfile

            with zipfile.ZipFile(self.BLANK_TEMPLATE_PATH, 'w', zipfile.ZIP_DEFLATED) as zf:
                content_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="no"?>
<xmap-content xmlns="urn:xmind:xmap:xmlns:content:2.0" xmlns:fo="http://www.w3.org/1999/XSL/Format" xmlns:svg="http://www.w3.org/2000/svg" xmlns:xhtml="http://www.w3.org/1999/xhtml" xmlns:xlink="http://www.w3.org/1999/xlink" modified-by="Lingma" timestamp="1234567890">
<sheet id="sheet1" modified-by="Lingma" timestamp="1234567890">
<topic id="root1" modified-by="Lingma" structure-class="org.xmind.ui.logic.right" timestamp="1234567890">
<title>Root</title>
</topic>
</sheet>
</xmap-content>'''
                zf.writestr('content.xml', content_xml.encode('utf-8'))

                meta_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="no"?>
<meta xmlns="urn:xmind:xmap:xmlns:meta:2.0" version="2.0">
<Creator>Lingma AI</Creator>
<Created>2024-01-01T00:00:00.000+0800</Created>
</meta>'''
                zf.writestr('meta.xml', meta_xml.encode('utf-8'))

            logger.info(f"✓ 已自动创建空白模板：{self.BLANK_TEMPLATE_PATH}")
        except Exception as e:
            logger.error(f"创建空白模板失败：{str(e)}")
            raise

    def _get_template_file(self, template_file=None):
        """获取有效的模板文件路径"""
        if template_file and os.path.exists(template_file):
            return template_file
        elif os.path.exists(self.BLANK_TEMPLATE_PATH):
            return self.BLANK_TEMPLATE_PATH
        elif os.path.exists(self.DEFAULT_TEMPLATE_PATH):
            return self.DEFAULT_TEMPLATE_PATH
        elif os.path.exists(self.FALLBACK_TEMPLATE_PATH):
            return self.FALLBACK_TEMPLATE_PATH
        else:
            raise FileNotFoundError(
                f"未找到 XMind 模板文件。\n"
                f"请确保以下路径之一存在：\n"
                f"1. {self.BLANK_TEMPLATE_PATH}（自动生成）\n"
                f"2. {self.DEFAULT_TEMPLATE_PATH}\n"
                f"3. {self.FALLBACK_TEMPLATE_PATH}"
            )

    def json_to_xmind(self, json_data, output_file, template_file=None):
        """
        将测试用例JSON数据转换为XMind文件

        Args:
            json_data: 测试用例JSON数据（字典或列表）
            output_file: 输出的XMind文件路径
            template_file: 可选的模板文件路径

        Returns:
            bool: 转换是否成功
        """
        try:
            valid_template = self._get_template_file(template_file)

            workbook = xmind.load(valid_template)
            sheet = workbook.getPrimarySheet()

            if not sheet:
                self.logger.error("模板文件中没有工作表")
                return False

            sheet_title = json_data.get('sheet_title', self.DEFAULT_SHEET_TITLE) if isinstance(json_data, dict) else self.DEFAULT_SHEET_TITLE
            sheet.setTitle(sheet_title)

            root_topic = sheet.getRootTopic()

            self._build_testcase_structure(root_topic, json_data)

            self._ensure_directory_exists(output_file)

            xmind.save(workbook, output_file)
            logger.info(f"XMind 文件已成功生成：{output_file}")
            return True

        except FileNotFoundError as e:
            logger.error(f"模板文件未找到：{str(e)}")
            return False
        except Exception as e:
            logger.error(f"转换 XMind 文件时发生错误：{str(e)}", exc_info=True)
            return False


    def _build_testcase_structure(self, root_topic, json_data):
        """
        构建测试用例的XMind结构
        支持多种JSON格式：
        1. {"测试用例": [...]}
        2. {"testcases": [...]}
        3. {"测试点名称": [{用例1}, {用例2}], ...}  ← 你的格式

        Args:
            root_topic: XMind根主题对象
            json_data: 测试用例JSON数据
        """
        if isinstance(json_data, dict):
            # 尝试标准格式：{"测试用例": [...]} 或 {"testcases": [...]}
            test_cases = json_data.get('测试用例', [])
            if not test_cases:
                test_cases = json_data.get('testcases', [])
            
            # 如果找到标准格式，按模块分组
            if test_cases and isinstance(test_cases, list):
                self._build_from_standard_format(root_topic, test_cases)
            else:
                # 否则，假设是 {"测试点名称": [用例列表]} 格式
                self._build_from_testpoint_format(root_topic, json_data)

        elif isinstance(json_data, list):
            # 直接是列表格式，按模块分组
            self._build_from_standard_format(root_topic, json_data)
        else:
            logger.warning(f"不支持的JSON数据类型: {type(json_data).__name__}")

    def _build_from_standard_format(self, root_topic, test_cases):
        """
        从标准格式构建XMind结构（按测试模块分组）
        
        Args:
            root_topic: XMind根主题对象
            test_cases: 测试用例列表
        """
        # 按模块分组 - 兼容多种字段名
        modules_dict = {}
        for item in test_cases:
            if isinstance(item, dict):
                # 兼容多种字段名："模块"、"测试模块"、"module"
                module_name = item.get('模块') or item.get('测试模块') or item.get('module', '未分类模块')
                
                # 获取用例列表 - 兼容多种字段名
                cases_list = item.get('用例') or item.get('testcases') or []
                
                if not cases_list:
                    logger.warning(f"模块 '{module_name}' 下没有找到用例列表")
                    continue
                
                if module_name not in modules_dict:
                    modules_dict[module_name] = []
                
                # 将该模块下的所有用例添加到列表
                for case in cases_list:
                    if isinstance(case, dict):
                        # 将模块名称注入到每个用例中，方便后续使用
                        case_copy = case.copy()
                        case_copy['测试模块'] = module_name
                        modules_dict[module_name].append(case_copy)

        # 为每个模块创建节点
        for module_name, cases in modules_dict.items():
            module_topic = root_topic.addSubTopic()
            module_topic.setTitle(module_name)

            # 为该模块下的每个用例创建子节点
            for testcase in cases:
                self._add_testcase_hierarchy(module_topic, testcase)

    def _build_from_testpoint_format(self, root_topic, json_data):
        """
        从测试点格式构建XMind结构（以测试点名称为分组）
        适用于格式：{"Z轴升降准确性": [{用例1}, {用例2}], ...}
        
        Args:
            root_topic: XMind根主题对象
            json_data: 测试用例JSON数据（字典，键为测试点名称）
        """
        for test_point_name, test_cases in json_data.items():
            if not isinstance(test_cases, list):
                logger.warning(f"测试点 '{test_point_name}' 的数据不是列表，跳过")
                continue
            
            # 创建测试点节点（作为一级分类）
            testpoint_topic = root_topic.addSubTopic()
            testpoint_topic.setTitle(test_point_name)
            
            # 为该测试点下的每个用例创建子节点
            for testcase in test_cases:
                if isinstance(testcase, dict):
                    self._add_testcase_hierarchy(testpoint_topic, testcase)


    def _add_testcase_hierarchy(self, parent_topic, testcase):
        """
        添加测试用例的完整层级结构（模块 -> 标题 -> 前置条件 -> 操作步骤 -> 预期结果）

        Args:
            parent_topic: 父主题对象（模块节点）
            testcase: 测试用例数据
        """
        # 第1层：用例标题作为模块的子节点
        title = testcase.get('标题', '未命名用例')
        priority = testcase.get('用例等级', '')

        title_topic = parent_topic.addSubTopic()
        title_topic.setTitle(title)

        # 添加优先级图标（使用 XMind 原生图标）
        if priority:
            # 将 P0, P1, P2, P3 等映射到 XMind 的优先级图标
            priority_map = {
                'P0': MarkerId.priority1,  # 优先级1（红色旗帜）
                'P1': MarkerId.priority2,  # 优先级2（橙色旗帜）
                'P2': MarkerId.priority3,  # 优先级3（黄色旗帜）
                'P3': MarkerId.priority4,  # 优先级4（绿色旗帜）
                'P4': MarkerId.priority5,  # 优先级5（蓝色旗帜）
                'P5': MarkerId.priority6,  # 优先级6（紫色旗帜）
                'P6': MarkerId.priority7,  # 优先级7（灰色旗帜）
                'P7': MarkerId.priority8,  # 优先级8（白色旗帜）
                'P8': MarkerId.priority9,  # 优先级9
            }

            marker_id = priority_map.get(priority)
            if marker_id:
                try:
                    title_topic.addMarker(marker_id)
                except Exception as e:
                    logger.warning(f"添加优先级图标失败: {str(e)}")

        # 第2层：前置条件作为标题的子节点
        precondition = testcase.get('前置条件', '')
        if precondition:
            precondition_topic = title_topic.addSubTopic()

            # 如果前置条件是列表，在当前节点用编号展示
            if isinstance(precondition, list):
                precondition_items = [f"{i+1}. {str(p)}" for i, p in enumerate(precondition)]
                precondition_topic.setTitle("\n".join(precondition_items))
            else:
                precondition_topic.setTitle(f"{precondition}")

            # 第3层：操作步骤作为前置条件的子节点
            steps = testcase.get('操作步骤', [])
            if steps:
                steps_topic = precondition_topic.addSubTopic()

                # 如果操作步骤是列表，在当前节点用编号展示
                if isinstance(steps, list):
                    steps_items = [f"{str(step)}" for i, step in enumerate(steps)]
                    steps_topic.setTitle("\n".join(steps_items))
                else:
                    steps_topic.setTitle(f"{steps}")

                # 第4层：预期结果作为操作步骤的子节点
                expected_result = testcase.get('预期结果', '')
                if expected_result:
                    result_topic = steps_topic.addSubTopic()

                    # 如果预期结果是列表，在当前节点用编号展示
                    if isinstance(expected_result, list):
                        result_items = [f"{str(r)}" for i, r in enumerate(expected_result)]
                        result_topic.setTitle("\n".join(result_items))
                    else:
                        result_topic.setTitle(f"{expected_result}")
            else:
                # 如果没有操作步骤，但有预期结果，预期结果作为前置条件的子节点
                expected_result = testcase.get('预期结果', '')
                if expected_result:
                    result_topic = precondition_topic.addSubTopic()

                    if isinstance(expected_result, list):
                        result_items = [f"{i+1}. {str(r)}" for i, r in enumerate(expected_result)]
                        result_topic.setTitle("预期结果：\n" + "\n".join(result_items))
                    else:
                        result_topic.setTitle(f"{expected_result}")
        else:
            # 如果没有前置条件，操作步骤作为标题的子节点
            steps = testcase.get('操作步骤', [])
            if steps:
                steps_topic = title_topic.addSubTopic()

                if isinstance(steps, list):
                    steps_items = [f"{str(step)}" for i, step in enumerate(steps)]
                    steps_topic.setTitle("\n".join(steps_items))
                else:
                    steps_topic.setTitle(f"{steps}")

                # 预期结果作为操作步骤的子节点
                expected_result = testcase.get('预期结果', '')
                if expected_result:
                    result_topic = steps_topic.addSubTopic()

                    if isinstance(expected_result, list):
                        result_items = [f"{str(r)}" for i, r in enumerate(expected_result)]
                        result_topic.setTitle("\n".join(result_items))
                    else:
                        result_topic.setTitle(f"{expected_result}")
            else:
                # 如果都没有，直接添加预期结果
                expected_result = testcase.get('预期结果', '')
                if expected_result:
                    result_topic = title_topic.addSubTopic()

                    if isinstance(expected_result, list):
                        result_items = [f"{str(r)}" for i, r in enumerate(expected_result)]
                        result_topic.setTitle("\n".join(result_items))
                    else:
                        result_topic.setTitle(f"{expected_result}")

    @staticmethod
    def extract_priority_from_title(title):
        """
        从标题中提取优先级标记

        Args:
            title: 标题字符串，如 "[P0] 测试模块1"

        Returns:
            tuple: (优先级, 清理后的标题)
                   例如: ("P0", "测试模块1")
        """
        # 匹配 [P0], [P1], [P2], [P3] 等格式
        pattern = r'^\[(P\d+)\]\s*(.+)$'
        match = re.match(pattern, title)

        if match:
            priority = match.group(1)
            clean_title = match.group(2)
            return priority, clean_title
        else:
            return None, title

    @staticmethod
    def add_priority_to_title(priority, title):
        """
        给标题添加优先级标记（已废弃，现在使用图标方式）

        Args:
            priority: 优先级，如 "P0", "P1"
            title: 原始标题

        Returns:
            str: 带优先级的标题（仅用于兼容旧代码）
        """
        if priority:
            return f"[{priority}] {title}"
        return title


    def convert_and_export_to_xmind(self, input_data, output_xmind_file,
                                    root_title=DEFAULT_ROOT_TITLE,
                                    sheet_title=DEFAULT_SHEET_TITLE,
                                    template_file=None):
        """
        一键转换并导出测试用例为XMind文件

        Args:
            input_data: 测试用例JSON数据
            output_xmind_file: 输出的XMind文件路径
            root_title: 根节点标题
            sheet_title: 工作表标题
            template_file: 可选的模板文件路径

        Returns:
            bool: 转换是否成功
        """
        try:
            if isinstance(input_data, dict):
                input_data['sheet_title'] = sheet_title

            success = self.json_to_xmind(input_data, output_xmind_file, template_file)

            if success:
                logger.info(f"XMind 文件已成功生成: {output_xmind_file}")

            return success

        except Exception as e:
            logger.error(f"转换过程中发生错误: {str(e)}", exc_info=True)
            return False

    @staticmethod
    def _ensure_directory_exists(file_path):
        """
        确保文件所在目录存在

        Args:
            file_path: 文件路径
        """
        file_dir = os.path.dirname(file_path)
        if file_dir and not os.path.exists(file_dir):
            os.makedirs(file_dir, exist_ok=True)

#测试用例的xmind格式文件转为相关格式json文件
class TestcaseXmindToAIJson:
    """
    读取xmind文件
    筛选过滤相关数据转化为相关格式json
    """

    def __init__(self):
        self.xmind_json_path = OUTPUT_JSON_PATH
        self.output_json_path = CONVERTED_TESTCASES_JSON_PATH
        self.logger = LogManager().get_logger() if hasattr(LogManager(), 'get_logger') else logger

    def read_xmind_json(self, json_file_path):
        """
        读取XMind格式的JSON文件
        Args:
            file_path: XMind JSON文件路径
        Returns:
            dict: XMind JSON数据
        """
        try:
            # 尝试用标准json.load读取
            try:
                with open(json_file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                self.logger.info(f"成功读取XMind JSON文件: {json_file_path}")
                return data
            except json.JSONDecodeError:
                # 如果失败，尝试用ast.literal_eval读取Python字典格式
                self.logger.info("检测到非标准JSON格式，尝试使用ast解析...")
                import ast
                with open(path, 'r', encoding='utf-8') as f:
                    content = f.read()

                # 将 None 替换为 null，True/False 替换为标准值
                content = content.replace('None', 'null').replace('True', 'true').replace('False', 'false')

                # 尝试再次用json加载
                try:
                    data = json.loads(content)
                    self.logger.info(f"成功读取XMind JSON文件（ast转换后）: {json_file_path}")
                    return data
                except:
                    # 最后尝试ast.literal_eval
                    data = ast.literal_eval(content)
                    self.logger.info(f"成功读取XMind JSON文件（ast解析）: {json_file_path}")
                    return data

        except Exception as e:
            self.logger.error(f"读取XMind JSON文件失败: {str(e)}")
            return None

    def extract_priority_from_marker(self, markers):
        """
        从标记中提取优先级
        Args:
            markers: 标记列表，如 ['priority-1']
        Returns:
            str: 优先级字符串，如 'P0'
        """
        if not markers:
            return ''

        priority_map = {
            'priority-1': 'P0',
            'priority-2': 'P1',
            'priority-3': 'P2',
            'priority-4': 'P3',
            'priority-5': 'P4',
            'priority-6': 'P5',
            'priority-7': 'P6',
            'priority-8': 'P7',
            'priority-9': 'P8',
        }

        for marker in markers:
            if marker in priority_map:
                return priority_map[marker]

        return ''

    def parse_steps_text(self, steps_text):
        """
        解析操作步骤文本，将其转换为列表
        Args:
            steps_text: 步骤文本，可能是多行字符串
        Returns:
            list: 步骤列表
        """
        if not steps_text:
            return []

        # 如果已经是列表，直接返回
        if isinstance(steps_text, list):
            return steps_text

        # 按换行符分割
        steps = [step.strip() for step in steps_text.split('\n') if step.strip()]
        return steps

    def traverse_and_extract(self, topics, current_module=None):
        """
        递归遍历XMind主题树，提取测试用例
        Args:
            topics: 主题列表
            current_module: 当前模块名称
        Returns:
            list: 测试用例列表
        """
        test_cases = []

        if not topics:
            return test_cases

        for topic in topics:
            title = topic.get('title', '')
            markers = topic.get('markers', [])
            sub_topics = topic.get('topics', [])

            # 如果没有子主题，跳过
            if not sub_topics:
                continue

            # 检查是否有优先级标记，如果有，说明这是测试用例节点
            has_priority = any(marker.startswith('priority-') for marker in markers)

            if has_priority:
                # 这是测试用例节点，提取完整信息
                testcase = self.extract_testcase_from_hierarchy(topic, current_module)
                if testcase:
                    test_cases.append(testcase)
            else:
                # 这可能是模块节点或其他中间节点，继续递归
                # 如果子主题有优先级标记，说明当前是模块节点
                first_sub = sub_topics[0] if sub_topics else None
                if first_sub and any(marker.startswith('priority-') for marker in first_sub.get('markers', [])):
                    # 当前是模块节点，设置模块名并递归处理子主题
                    module_name = title
                    for sub_topic in sub_topics:
                        sub_cases = self.traverse_and_extract([sub_topic], current_module=module_name)
                        test_cases.extend(sub_cases)
                else:
                    # 继续递归查找
                    sub_cases = self.traverse_and_extract(sub_topics, current_module=current_module)
                    test_cases.extend(sub_cases)

        return test_cases

    def extract_testcase_from_hierarchy(self, topic_node, module_name):
        """
        从层级结构中提取单个测试用例
        Args:
            topic_node: 测试用例主题节点
            module_name: 所属模块名称
        Returns:
            dict: 测试用例字典
        """
        try:
            # 第1层：测试用例标题
            title = topic_node.get('title', '')
            markers = topic_node.get('markers', [])
            priority = self.extract_priority_from_marker(markers)

            # 获取子主题
            level1_topics = topic_node.get('topics', [])
            if not level1_topics:
                return None

            # 第2层：前置条件
            precondition_topic = level1_topics[0]
            precondition = precondition_topic.get('title', '')

            # 获取前置条件的子主题
            level2_topics = precondition_topic.get('topics', [])
            if not level2_topics:
                return None

            # 第3层：操作步骤
            steps_topic = level2_topics[0]
            steps_text = steps_topic.get('title', '')
            steps = self.parse_steps_text(steps_text)

            # 获取操作步骤的子主题
            level3_topics = steps_topic.get('topics', [])
            expected_result = ''
            if level3_topics:
                # 第4层：预期结果
                result_topic = level3_topics[0]
                expected_result = result_topic.get('title', '')

            # 构建测试用例对象
            testcase = {
                '测试模块': module_name or '',
                '用例等级': priority,
                '标题': title,
                '前置条件': precondition,
                '操作步骤': steps,
                '预期结果': expected_result
            }

            return testcase

        except Exception as e:
            self.logger.warning(f"提取测试用例失败: {str(e)}")
            return None

    def convert_to_testcase_format(self, xmind_data):
        """
        将XMind格式数据转换为测试用例JSON格式
        Args:
            xmind_data: XMind JSON数据
        Returns:
            dict: 测试用例JSON格式数据
        """
        try:
            # 获取根主题
            root_topic = xmind_data.get('topic', {})
            if not root_topic:
                self.logger.error("XMind数据中没有找到topic节点")
                return {'测试用例': []}

            # 获取第一层主题（模块级别）
            modules = root_topic.get('topics', [])

            # 遍历并提取所有测试用例
            all_testcases = []
            for module_topic in modules:
                module_name = module_topic.get('title', '')
                sub_topics = module_topic.get('topics', [])

                # 对每个模块下的用例进行提取
                module_testcases = self.traverse_and_extract([module_topic])
                all_testcases.extend(module_testcases)

            # 构建最终结果
            result = {
                '测试用例': all_testcases
            }

            self.logger.info(f"成功转换 {len(all_testcases)} 条测试用例")
            return result

        except Exception as e:
            self.logger.error(f"转换数据格式失败: {str(e)}", exc_info=True)
            return {'测试用例': []}

    def save_to_json(self, data, output_path=None):
        """
        保存为JSON文件
        Args:
            data: 要保存的数据
            output_path: 输出文件路径
        Returns:
            bool: 是否保存成功
        """
        try:
            path = output_path or self.output_json_path

            # 确保目录存在
            output_dir = os.path.dirname(path)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)

            with open(path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            self.logger.info(f"测试用例JSON已保存到: {path}")
            print(f"✓ 测试用例JSON已保存到: {path}")
            return True

        except Exception as e:
            self.logger.error(f"保存JSON文件失败: {str(e)}")
            print(f"✗ 保存失败: {str(e)}")
            return False

    def convert(self, input_path=None, output_path=None):
        """
        主转换方法：读取XMind JSON并转换为测试用例JSON
        Args:
            input_path: 输入XMind JSON文件路径
            output_path: 输出测试用例JSON文件路径
        Returns:
            dict: 转换后的测试用例数据
        """
        print("开始转换XMind数据为测试用例JSON...")

        # 1. 读取XMind JSON
        xmind_data = self.read_xmind_json(input_path)
        if not xmind_data:
            print("✗ 读取XMind数据失败")
            return None

        # 2. 转换为测试用例格式
        testcase_data = self.convert_to_testcase_format(xmind_data)

        # 3. 保存结果
        success = self.save_to_json(testcase_data, output_path)

        if success:
            print(f"✓ 转换完成！共生成 {len(testcase_data.get('测试用例', []))} 条测试用例")
        else:
            print("✗ 转换失败")

        return testcase_data

#AI给出json，转化为execl格式文件
class AIJsonToXlsx:

    def __init__(self):
        self.file_path = TESTCASE_JSON_PATH
        self.output_file = OUTPUT_XLSX_PATH
        self.sheet_name = "Sheet1"
        self.column_names = []

    # 读取数据
    def read_data(self):
        get_json = fileProcessor()
        test_data = get_json.find_and_read_file(self.file_path, type="json")
        print(isinstance(test_data, dict))
        return test_data

    # 过滤筛选
    def filter_data(self):
        """
        过滤和整理测试用例数据
        Returns:
            list: 整理后的测试用例列表
        """
        test_data = self.read_data()
        if not test_data or '测试用例' not in test_data:
            print("未找到有效的测试用例数据")
            return []
        
        test_cases = test_data['测试用例']
        filtered_cases = []
        
        for case in test_cases:
            if isinstance(case, dict):
                # 确保必要的字段存在
                filtered_case = {
                    '测试模块': case.get('测试模块', ''),
                    '用例等级': case.get('用例等级', ''),
                    '标题': case.get('标题', ''),
                    '前置条件': case.get('前置条件', ''),
                    '操作步骤': case.get('操作步骤', []),
                    '预期结果': case.get('预期结果', '')
                }
                filtered_cases.append(filtered_case)
        
        return filtered_cases

    # 创建模板  自定义字段+ 字段的键值字段  模板参考tapd的标准
    def create_template(self):
        """
        创建Excel模板，定义表头结构
        Returns:
            list: 表头列表
        """
        headers = ['测试模块', '用例等级', '标题', '前置条件', '操作步骤', '预期结果']
        self.column_names = headers
        return headers

    # 写入数据
    def write_data(self, test_cases):
        """
        将测试用例数据写入Excel文件
        Args:
            test_cases: 测试用例列表
        """
        try:
            # 创建工作簿和工作表
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            
            # 定义表头样式：绿色背景 + 粗体字体
            header_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            header_font = Font(bold=True)
            
            # 写入表头
            headers = self.create_template()
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # 设置表头行高为25
            ws.row_dimensions[1].height = 25
            
            # 写入数据行
            for row_num, case in enumerate(test_cases, 2):  # 从第2行开始写数据
                # 处理操作步骤（可能是列表或字符串）
                steps = case.get('操作步骤', [])
                if isinstance(steps, list):
                    steps_text = '\n'.join([str(step) for step in steps])
                else:
                    steps_text = str(steps)
                
                # 写入各列数据
                ws.cell(row=row_num, column=1, value=case.get('测试模块', ''))
                ws.cell(row=row_num, column=2, value=case.get('用例等级', ''))
                ws.cell(row=row_num, column=3, value=case.get('标题', ''))
                ws.cell(row=row_num, column=4, value=case.get('前置条件', ''))
                ws.cell(row=row_num, column=5, value=steps_text)
                ws.cell(row=row_num, column=6, value=case.get('预期结果', ''))
                
                # 设置所有数据行的行高为25
                ws.row_dimensions[row_num].height = 25
            
            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # 最大宽度限制为50
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存文件
            wb.save(self.output_file)
            print(f"测试用例已成功导出到: {self.output_file}")
            return True
            
        except Exception as e:
            print(f"写入Excel文件时发生错误: {str(e)}")
            return False

    # 导出
    def export(self):
        """
        主导出方法，执行完整的数据转换流程
        Returns:
            bool: 导出是否成功
        """
        print("开始转换测试用例数据...")
        
        # 1. 读取并过滤数据
        test_cases = self.filter_data()
        if not test_cases:
            print("没有可导出的测试用例数据")
            return False
        
        print(f"共找到 {len(test_cases)} 条测试用例")
        
        # 2. 写入Excel
        success = self.write_data(test_cases)
        
        if success:
            print("测试用例转换完成！")
        else:
            print("测试用例转换失败！")
        
        return success

#测试点json提取成符合测试用例输入规则标准的data,然后写入到输出的json文件中
class TestPointToAIJson:
    """
    将测试点数据转换为简化的测试转化数据格式
    Attributes:
        logger: 日志记录器
    """
    def __init__(self):
        """
        初始化转换器
        """

    def convert_testpoint_to_simple_format(self, input_data: dict) -> dict:
        """
        将测试点数据转换为简化的测试转化数据格式
        提取测试模块下的测试点名称和场景描述，重组为扁平化结构
        
        Args:
            input_data: 测试点JSON数据
            
        Returns:
            Dict: 转换后的简化格式数据
            
        Raises:
            TypeError: 当输入数据类型不正确时抛出
            ValueError: 当输入数据为空时抛出
        """
        if not input_data:
            raise ValueError("输入数据不能为空")
        
        if not isinstance(input_data, dict):
            raise TypeError(f"期望输入类型为dict，实际得到: {type(input_data).__name__}")
        
        result = {}
        
        # 遍历顶层产品/需求名称
        for product_name, modules_list in input_data.items():
            logger.info(f"正在处理产品: {product_name}")
            
            if not isinstance(modules_list, list):
                logger.warning(f"产品 '{product_name}' 的数据不是列表格式，跳过")
                continue
            
            # 遍历每个功能模块
            for module_item in modules_list:
                if not isinstance(module_item, dict):
                    continue
                
                for module_name, test_points_list in module_item.items():
                    logger.debug(f"正在处理模块: {module_name}")
                    
                    # 遍历测试点列表
                    if isinstance(test_points_list, list):
                        for test_point_item in test_points_list:
                            if not isinstance(test_point_item, dict):
                                continue
                            
                            for test_point_name, scenarios_list in test_point_item.items():
                                # 提取该测试点下的所有场景描述
                                scenario_descriptions = []
                                
                                if isinstance(scenarios_list, list):
                                    for scenario in scenarios_list:
                                        if isinstance(scenario, dict):
                                            # 提取场景描述
                                            desc = self._extract_scenario_description(scenario)
                                            if desc:
                                                scenario_descriptions.append(desc)
                                
                                # 只有当有场景描述时才添加到结果中
                                if scenario_descriptions:
                                    result[test_point_name] = scenario_descriptions
                                    logger.debug(
                                        f"提取测试点 '{test_point_name}': {len(scenario_descriptions)} 个场景"
                                    )
        
        logger.info(
            f"转换完成，共提取 {len(result)} 个测试点"
        )
        return result

    def _extract_scenario_description(self, scenario: dict) -> str:
        """
        从场景字典中提取格式化的场景描述
        
        Args:
            scenario: 场景字典，如 {"正向：Z轴升降至指定高度位置精度准确": "输入目标高度，Z轴准确移动到对应位置"}
            
        Returns:
            str: 格式化后的场景描述，如 "正向-Z轴升降至指定高度位置精度准确"
        """
        if not scenario or not isinstance(scenario, dict):
            return ""
        
        # 获取第一个键值对
        for key, value in scenario.items():
            # 判断是否包含中文冒号或英文冒号
            if "：" in key:
                parts = key.split("：", 1)
                scenario_type = parts[0]  # 正向/反向
                scenario_title = parts[1] if len(parts) > 1 else ""  # 场景标题
            elif ":" in key:
                parts = key.split(":", 1)
                scenario_type = parts[0]
                scenario_title = parts[1] if len(parts) > 1 else ""
            else:
                # 没有分隔符，直接使用整个key
                scenario_type = ""
                scenario_title = key
            
            # 组合格式：类型-标题
            if scenario_type and scenario_title:
                return f"{scenario_type}-{scenario_title}"
            elif scenario_title:
                return scenario_title
            else:
                return key
        
        return ""

    def save_to_json(self, data: dict, output_file: str) -> str:
        """
        将转换后的数据保存为JSON文件
        
        Args:
            data: 要保存的JSON数据
            output_file: 输出文件路径
            
        Returns:
            str: 保存文件的绝对路径
            
        Raises:
            ValueError: 当输出文件路径为空时抛出
            IOError: 当文件写入失败时抛出
        """
        if not output_file:
            raise ValueError("输出文件路径不能为空")
        
        # 确保输出目录存在
        output_dir = os.path.dirname(output_file)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
            logger.info(f"创建输出目录: {output_dir}")
        
        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
            
            abs_path = os.path.abspath(output_file)
            logger.info(f"JSON文件保存成功: {abs_path}")
            return abs_path
            
        except IOError as e:
            logger.error(f"文件写入失败: {output_file}, 错误: {str(e)}")
            raise

    def convert_and_save(self, input_file: str, output_file: str,extract: bool) -> str:
        """
        读取测试点JSON文件，转换为简化格式并保存
        
        Args:
            input_file: 输入的测试点JSON文件路径
            output_file: 输出的简化格式JSON文件路径
            
        Returns:
            str: 输出文件的绝对路径
            
        Raises:
            FileNotFoundError: 当输入文件不存在时抛出
            json.JSONDecodeError: 当JSON解析失败时抛出
        """
        # 读取输入文件
        if not os.path.exists(input_file):
            raise FileNotFoundError(f"输入文件不存在: {input_file}")
        
        logger.info(f"开始读取测试点文件: {input_file}")
        
        try:
            with open(input_file, 'r', encoding='utf-8') as f:
                input_data = json.load(f)
        except json.JSONDecodeError as e:
            logger.error(f"JSON解析失败: {str(e)}")
            raise
        
        # 转换数据

        logger.info("开始转换数据...")
        res_data = self.convert_testpoint_to_simple_format(input_data)
        if extract:
            # 保存结果
            logger.info(f"保存转换结果到: {output_file}")
            output_path = self.save_to_json(res_data, output_file)
            # 统计信息
            test_point_count = len(res_data)
            scenario_count = sum(len(scenarios) for scenarios in res_data.values())

            logger.info(
                f"转换完成！共 {test_point_count} 个测试点，{scenario_count} 个场景"
            )
        else:
            logger.info("不保存结果到文件中")
        return res_data


#图片识别文字，AI过滤处理
class TextRecognition:
    """文本识别与AI过滤工具类"""
    
    def __init__(self, image_path=None):
        """
        初始化文本识别器
        
        Args:
            image_path: 图片路径，默认为配置文件中的IMG_PATH
        """
        self.ocr_img = image_path if image_path else IMG_PATH
        logger.info(f"TextRecognition初始化完成，图片路径: {self.ocr_img}")
    
    def get_ai_point(self, user_input):
        """
        公开方法：从图片识别文字并通过AI过滤获取测试点
        
        Args:
            user_input: 用户输入的需求描述
            
        Returns:
            str: AI过滤后的测试结果
            
        Raises:
            ValueError: 当参数为空或无效时抛出
            FileNotFoundError: 当图片文件不存在时抛出
        """
        # 入参校验
        if not user_input or not isinstance(user_input, str):
            logger.error("user_input参数不能为空且必须为字符串类型")
            raise ValueError("user_input参数不能为空且必须为字符串类型")
        
        if not self.ocr_img or not isinstance(self.ocr_img, str):
            logger.error("图片路径不能为空且必须为字符串类型")
            raise ValueError("图片路径不能为空且必须为字符串类型")
        
        try:
            logger.info(f"开始处理用户输入: {user_input[:50]}...")
            
            # 步骤1：从图片中提取文字（直接调用公共函数）
            ocr_result = ocr_get_text_from_img(self.ocr_img)
            logger.debug(f"OCR识别完成，共识别到 {len(ocr_result.get('text', []))} 段文字")
            
            # 步骤2：通过AI过滤处理
            filtered_result = self._text_to_filter_ai(ocr_result, user_input)
            
            logger.info("AI过滤处理完成")
            return filtered_result
            
        except FileNotFoundError as e:
            logger.error(f"图片文件不存在: {str(e)}")
            raise
        except Exception as e:
            logger.error(f"处理过程中发生错误: {str(e)}", exc_info=True)
            raise
    
    def _text_to_filter_ai(self, ocr_result: dict, user_input: str) -> str:
        """
        私有方法：将OCR识别的文字通过AI进行过滤处理
        
        Args:
            ocr_result: OCR识别结果字典
            user_input: 用户输入的需求描述
            
        Returns:
            str: AI处理后的结果
        """
        logger.debug("开始AI过滤处理")
        
        try:
            # 提取OCR识别的文本内容
            text_list = ocr_result.get('text', [])
            if not text_list:
                logger.warning("OCR未识别到任何文字内容")
                return ""
            
            # 将识别的文字列表拼接为字符串
            ocr_text = "\n".join(text_list)
            logger.debug(f"待处理的OCR文本长度: {len(ocr_text)} 字符")
            
            # 调用DeepSeek API进行处理
            ai_api = DeepSeekAPI()
            ai_result = ai_api.get_ai_point(user_input)
            
            logger.info("AI处理成功")
            return ai_result
            
        except Exception as e:
            logger.error(f"AI过滤处理失败: {str(e)}", exc_info=True)
            raise




if __name__ == '__main__':
    # converter = TestPointToAIJson()
    #
    # # 一键转换并保存
    # output_path = converter.convert_and_save(
    #     input_file=r'D:\AIGeneration\testcase\测试点.json',
    #     output_file=r'D:\AIGeneration\testcase\测试转化数据.json'
    # )
    #
    # print(f"转换成功！输出文件: {output_path}")
    #
    # converter = XmindPointJson(xmind_file=r"C:/Users/admin/Desktop/demo.xmind")
    # result = converter.process_and_save(output_file=r"D:/AIGeneration/testcase/output.json")






    # dc = MxindDataProcessor()
    # res = dc.write_to_json()

    # #
    # converter = TestcaseXmindToAIJson()
    #
    # # 步骤1：读取XMind JSON
    # xmind_data = converter.read_xmind_json(r"D:\AIGeneration\testcase\xmind_output.json")
    #
    # # 步骤2：转换为测试用例格式
    # testcase_data = converter.convert_to_testcase_format(xmind_data)
    #
    # # 步骤3：保存
    # converter.save_to_json(testcase_data, r"D:\AIGeneration\testcase\output.json")
    #



    # converter = AIJsonToXlsx()
    #
    # # 步骤1：读取数据
    # data = converter.read_data()
    # print(data)
    #
    # # 步骤2：过滤数据
    # test_cases = converter.filter_data()
    # print(f"共 {len(test_cases)} 条用例")
    #
    # # 步骤3：写入Excel
    # success = converter.write_data(test_cases)
    # if success:
    #     print("转换成功！")


    # dic_to_xlsx = DicToXlsx()
    # dic_to_xlsx.table_data_processing()

    # DD = WriteInfo()
    # DD.write_json_to_file(data=res, file=r"D:\AIGeneration\testcase\demo.json")
    #
    # with open(r"D:\AIGeneration\testcase\demo.json", 'r', encoding='utf-8') as f:
    #     json_data = json.load(f)
    # # # 转换为 XMind并导出
    # output_xmind = r"C:\Users\admin\Desktop\demo_output.xmind"
    # json_to_xmind(json_data, output_xmind)

    # res = XmindPointJson()
    # res.extract_test_points_data()

    #
    #
    # with open(r'D:\AIGeneration\testcase\测试点.json', 'r', encoding='utf-8') as f:
    #     data = json.load(f)
    #
    # # 2. 实例化工具类
    # converter = TestcasePointJsonToXmind()
    #
    # # 3. 执行转换并导出 XMind
    # # root_title: 整个思维导图的最外层根节点名称
    # # sheet_title: 工作表的标题
    # success = converter.convert_and_export_to_xmind(
    #     input_data=data,
    #     output_xmind_file=r'D:\AIGeneration\testcase\output_test_points.xmind',
    #     root_title="功能测试",
    #     sheet_title="LCD打印机测试点"
    # )
    #
    # if success:
    #     print("XMind 文件生成成功！")
    # else:
    #     print("生成失败，请检查日志。")
    #




    # 测试用例JSON转XMind示例
    print("\n=== 测试用例JSON转XMind ===")
    testcase_converter = TestcaseJsonToXmind()

    # 读取测试用例JSON文件
    testcase_json_path = r"D:\AIGeneration\testcase\测试用例_output.json"
    get_json = fileProcessor()
    testcase_data = get_json.find_and_read_file(testcase_json_path, type="json")

    # 转换并导出
    output_xmind = r"D:\AIGeneration\testcase\测试用例.xmind"
    success = testcase_converter.convert_and_export_to_xmind(
        input_data=testcase_data,
        output_xmind_file=output_xmind,
        root_title="测试用例",
        sheet_title="功能测试用例"
    )

    if success:
        print(f"✓ 测试用例XMind文件已生成：{output_xmind}")
    else:
        print("✗ 转换失败")